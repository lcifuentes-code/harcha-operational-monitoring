"""
app.py  —  Dashboard ejecutivo de monitoreo operacional
=========================================================
Streamlit · Harcha Maquinaria

Ejecutar:
    streamlit run app.py

Solo este archivo cambió respecto a la versión anterior.
Los scripts de procesamiento (scripts/) no se modificaron.
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from scripts.load_data  import cargar_datos
from scripts.clean_data import limpiar_datos
from scripts.transform  import calcular_metricas
from scripts.alerts     import generar_alertas
from config.settings    import ESTADOS_ACTIVOS

# ── Configuración de página ───────────────────────────────────────────────────
st.set_page_config(
    page_title="Harcha · Monitor Operacional",
    page_icon="🚜",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── CSS global del dashboard ──────────────────────────────────────────────────
st.markdown("""
<style>
:root {
    --azul:        #3B82F6;
    --azul-oscuro: #1D4ED8;
    --azul-claro:  #EFF6FF;
    --rojo:        #EF4444;
    --rojo-claro:  #FEF2F2;
    --verde:       #10B981;
    --verde-claro: #ECFDF5;
    --ambar:       #F59E0B;
    --ambar-claro: #FFFBEB;
    --gris-bg:     #F8FAFC;
    --gris-borde:  #E2E8F0;
    --texto-1:     #0F172A;
    --texto-2:     #475569;
    --texto-3:     #94A3B8;
    --blanco:      #FFFFFF;
    --radio:       12px;
    --sombra:      0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.04);
    --sombra-md:   0 4px 12px rgba(0,0,0,.08);
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 1.5rem 2rem 2rem !important; max-width: 1400px !important; }

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px; background: var(--gris-bg); padding: 5px;
    border-radius: 10px; border: 1px solid var(--gris-borde);
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px !important; padding: 8px 20px !important;
    font-size: 14px; font-weight: 500; color: var(--texto-2);
    background: transparent; border: none !important;
}
.stTabs [aria-selected="true"] {
    background: var(--blanco) !important; color: var(--azul) !important;
    box-shadow: var(--sombra) !important;
}
.stTabs [data-baseweb="tab-border"] { display: none; }
.stTabs [data-baseweb="tab-panel"] { padding-top: 1.25rem !important; }

/* Ocultar colapso del sidebar */
div[data-testid="stSidebarCollapsedControl"] button { display: none; }

/* KPI Grid */
.kpi-grid { display: grid; grid-template-columns: repeat(5, 1fr); gap: 14px; margin-bottom: 20px; }
.kpi-card {
    background: var(--blanco); border: 1px solid var(--gris-borde);
    border-radius: var(--radio); padding: 18px 20px 16px;
    box-shadow: var(--sombra); position: relative; overflow: hidden;
    transition: box-shadow .15s, transform .15s;
}
.kpi-card:hover { box-shadow: var(--sombra-md); transform: translateY(-1px); }
.kpi-card::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0;
    height: 3px; background: var(--kpi-color, var(--azul));
    border-radius: var(--radio) var(--radio) 0 0;
}
.kpi-icon  { font-size: 20px; margin-bottom: 8px; display: block; }
.kpi-label { font-size: 11px; font-weight: 700; letter-spacing: .07em;
             text-transform: uppercase; color: var(--texto-3); margin-bottom: 5px; }
.kpi-value { font-size: 28px; font-weight: 800; color: var(--texto-1);
             line-height: 1; margin-bottom: 4px; }
.kpi-sub   { font-size: 11.5px; color: var(--texto-2); }
.kpi-alerta { --kpi-color: var(--rojo); }
.kpi-verde  { --kpi-color: var(--verde); }
.kpi-ambar  { --kpi-color: var(--ambar); }

/* Sección / card interna */
.seccion {
    background: var(--blanco); border: 1px solid var(--gris-borde);
    border-radius: var(--radio); padding: 18px 20px;
    box-shadow: var(--sombra); margin-bottom: 16px;
}
.seccion-titulo {
    font-size: 13.5px; font-weight: 700; color: var(--texto-1);
    margin: 0 0 14px; display: flex; align-items: center; gap: 6px;
}

/* Badge */
.badge {
    display: inline-block; padding: 2px 9px; border-radius: 20px;
    font-size: 10.5px; font-weight: 700; letter-spacing: .05em;
}
.badge-azul  { background: var(--azul-claro);  color: var(--azul-oscuro); }
.badge-rojo  { background: var(--rojo-claro);  color: #B91C1C; }
.badge-verde { background: var(--verde-claro); color: #065F46; }
.badge-ambar { background: var(--ambar-claro); color: #92400E; }

/* Alerta item */
.alerta-item {
    display: flex; align-items: flex-start; gap: 10px;
    padding: 11px 13px; border-radius: 8px; margin-bottom: 8px; border: 1px solid;
}
.alerta-rojo  { background: var(--rojo-claro);  border-color: #FECACA; }
.alerta-ambar { background: var(--ambar-claro); border-color: #FDE68A; }
.alerta-dot   { width: 7px; height: 7px; border-radius: 50%;
                flex-shrink: 0; margin-top: 5px; }
.alerta-dot-r { background: var(--rojo); }
.alerta-dot-a { background: var(--ambar); }
.alerta-maq   { font-size: 13px; font-weight: 600; color: var(--texto-1); margin-bottom: 2px; }
.alerta-desc  { font-size: 11.5px; color: var(--texto-2); }

/* Ranking */
.rank-row {
    display: flex; align-items: center; padding: 9px 0;
    border-bottom: 1px solid var(--gris-borde); gap: 10px;
}
.rank-row:last-child { border-bottom: none; }
.rank-pos    { font-size: 13px; font-weight: 700; color: var(--texto-3); width: 22px; text-align: center; }
.rank-nombre { font-size: 13px; font-weight: 600; color: var(--texto-1); flex: 1; }
.rank-valor  { font-size: 13px; font-weight: 700; color: var(--azul); }
.rank-sub    { font-size: 11px; color: var(--texto-3); }

/* Barra de progreso */
.prog-wrap   { margin-bottom: 11px; }
.prog-header { display: flex; justify-content: space-between;
               font-size: 12px; margin-bottom: 4px; }
.prog-label  { font-weight: 600; color: var(--texto-1); }
.prog-val    { color: var(--texto-2); }
.prog-bar-bg { height: 5px; background: var(--gris-borde);
               border-radius: 4px; overflow: hidden; }
.prog-bar-fill { height: 100%; border-radius: 4px; }

/* Upload zone */
.upload-zone {
    background: var(--azul-claro); border: 2px dashed #93C5FD;
    border-radius: var(--radio); padding: 32px; text-align: center; margin: 20px 0;
}
.upload-title { font-size: 18px; font-weight: 700; color: var(--azul-oscuro); margin-bottom: 6px; }
.upload-sub   { font-size: 13px; color: var(--texto-2); }

/* Tablas */
div[data-testid="stDataFrame"] > div {
    border-radius: 8px !important; border: 1px solid var(--gris-borde) !important;
}
</style>
""", unsafe_allow_html=True)


# ── Helpers de render ─────────────────────────────────────────────────────────
def kpi_card(icono, label, valor, sub="", color=""):
    clase = f"kpi-card {color}"
    return f"""
    <div class="{clase}">
        <span class="kpi-icon">{icono}</span>
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{valor}</div>
        <div class="kpi-sub">{sub}</div>
    </div>"""


def barra(label, valor, maximo, color="#3B82F6", sub=""):
    pct = min(100, round((valor / maximo * 100) if maximo > 0 else 0, 1))
    return f"""
    <div class="prog-wrap">
        <div class="prog-header">
            <span class="prog-label">{label}</span>
            <span class="prog-val">{sub}</span>
        </div>
        <div class="prog-bar-bg">
            <div class="prog-bar-fill" style="width:{pct}%; background:{color}"></div>
        </div>
    </div>"""


# ── Cache de procesamiento ────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def procesar(bytes_archivo: bytes, nombre: str) -> dict:
    buf = io.BytesIO(bytes_archivo)
    buf.name = nombre
    datos    = limpiar_datos(cargar_datos(archivo=buf))
    metricas = calcular_metricas(datos)
    alertas  = generar_alertas(datos)
    return {"datos": datos, "metricas": metricas, "alertas": alertas}


# ── Pantalla de carga ─────────────────────────────────────────────────────────
def pantalla_inicio():
    st.markdown("""
    <div style="text-align:center; padding:60px 0 20px">
        <div style="font-size:52px; margin-bottom:12px">🚜</div>
        <div style="font-size:28px; font-weight:800; color:#0F172A; margin-bottom:8px">
            Monitor Operacional
        </div>
        <div style="font-size:15px; color:#64748B; margin-bottom:32px">
            Harcha Maquinaria · Dashboard ejecutivo
        </div>
    </div>
    """, unsafe_allow_html=True)

    _, col_cen, _ = st.columns([1, 2, 1])
    with col_cen:
        st.markdown("""
        <div class="upload-zone">
            <div class="upload-title">📂 Sube el archivo Excel</div>
            <div class="upload-sub">Descarga el reporte desde Google Sheets y súbelo aquí</div>
        </div>
        """, unsafe_allow_html=True)
        archivo = st.file_uploader("", type=["xlsx"], label_visibility="collapsed")
        if archivo:
            return archivo

        st.markdown("""
        <div style="margin-top:20px; background:#F8FAFC; border:1px solid #E2E8F0;
                    border-radius:12px; padding:18px 22px">
            <div style="font-size:13px; font-weight:700; color:#0F172A; margin-bottom:8px">
                ¿Cómo obtener el archivo?
            </div>
            <div style="font-size:13px; color:#475569; line-height:1.9">
                1. Abre Google Sheets desde AppSheet<br>
                2. Archivo → Descargar → Microsoft Excel (.xlsx)<br>
                3. Sube el archivo descargado aquí
            </div>
        </div>

        <!-- Preparado para Google Sheets directo (fase futura) →
        Descomentar cuando se habilite la conexión directa:

        gsheet_url = st.text_input("URL Google Sheets", placeholder="https://docs.google.com/...")
        import gspread
        from google.oauth2 import service_account
        creds = service_account.Credentials.from_service_account_file("credentials.json")
        gc = gspread.authorize(creds)
        ...
        -->
        """, unsafe_allow_html=True)
    return None


# ── Estado de sesión ──────────────────────────────────────────────────────────
if "archivo_bytes" not in st.session_state:
    res = pantalla_inicio()
    if res is None:
        st.stop()
    st.session_state.archivo_bytes  = res.read()
    st.session_state.archivo_nombre = res.name
    st.rerun()

# ── Procesar datos ────────────────────────────────────────────────────────────
with st.spinner("Procesando datos..."):
    resultado = procesar(st.session_state.archivo_bytes, st.session_state.archivo_nombre)

datos         = resultado["datos"]
alertas_base  = resultado["alertas"]
df_rep_full   = datos["reportes"]

# ── BARRA SUPERIOR: logo · filtros · cambiar archivo ─────────────────────────
col_logo, col_f1, col_f2, col_btn = st.columns([2, 2, 2, 1.5])

with col_logo:
    st.markdown("""
    <div style="display:flex;align-items:center;gap:10px;padding:8px 0">
        <span style="font-size:28px">🚜</span>
        <div>
            <div style="font-size:15px;font-weight:800;color:#0F172A;line-height:1.1">
                Harcha Monitor
            </div>
            <div style="font-size:11px;color:#94A3B8">Dashboard Operacional</div>
        </div>
    </div>""", unsafe_allow_html=True)

fecha_min = df_rep_full["FECHAHORA_INICIO"].min().date()
fecha_max = df_rep_full["FECHAHORA_INICIO"].max().date()

with col_f1:
    fecha_ini = st.date_input("📅 Desde", value=fecha_min,
                               min_value=fecha_min, max_value=fecha_max)
with col_f2:
    fecha_fin = st.date_input("📅 Hasta", value=fecha_max,
                               min_value=fecha_min, max_value=fecha_max)
with col_btn:
    st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
    if st.button("🔄 Cambiar archivo", use_container_width=True):
        for k in ["archivo_bytes", "archivo_nombre"]:
            del st.session_state[k]
        st.rerun()

st.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)

# ── Aplicar filtro de fechas ──────────────────────────────────────────────────
if fecha_ini > fecha_fin:
    st.error("⚠️ La fecha inicial no puede ser mayor a la final.")
    st.stop()

df_rep = df_rep_full[
    (df_rep_full["FECHAHORA_INICIO"].dt.date >= fecha_ini) &
    (df_rep_full["FECHAHORA_INICIO"].dt.date <= fecha_fin)
].copy()

datos_filtrados = {**datos, "reportes": df_rep}
metricas = calcular_metricas(datos_filtrados)
alertas  = generar_alertas(datos_filtrados)

maq_df  = metricas.get("por_maquina",     pd.DataFrame())
ops_df  = metricas.get("operadores",      pd.DataFrame())
comb_df = metricas.get("combustible",     pd.DataFrame())
act_df  = metricas.get("actividad_diaria",pd.DataFrame())

total_horas  = maq_df["total_horas"].sum()   if not maq_df.empty  else 0
total_litros = comb_df["total_litros"].sum()  if not comb_df.empty else 0
n_maquinas   = maq_df["ID_MAQUINA"].nunique() if not maq_df.empty  else 0
n_operadores = ops_df["USUARIO_TXT"].nunique()if not ops_df.empty  else 0
n_alertas    = len(alertas) if not alertas.empty else 0
n_dias       = (fecha_fin - fecha_ini).days + 1

# ── PESTAÑAS ─────────────────────────────────────────────────────────────────
tab_act, tab_maq, tab_ops, tab_comb, tab_alertas = st.tabs([
    "🎯  Vista Ejecutiva",
    "🚜  Máquinas",
    "👷  Operadores",
    "⛽  Combustible",
    f"🚨  Alertas  ({n_alertas})",
])



# ════════════════════════════════════════════════════════════
# TAB 1 — VISTA EJECUTIVA
# ════════════════════════════════════════════════════════════
with tab_act:

    # ── Cálculos centrales ────────────────────────────────────────────────────

    df_maq_cat       = datos.get("maquinas", pd.DataFrame())
    total_produccion = len(df_maq_cat[df_maq_cat["ESTADO"].isin(ESTADOS_ACTIVOS)]) \
                       if not df_maq_cat.empty else max(n_maquinas, 1)
    salud_pct        = min(round((n_maquinas / total_produccion * 100)
                                 if total_produccion > 0 else 0, 1), 100.0)

    # % cumplimiento de reporte (máquinas en producción que reportaron al menos 1 vez)
    maquinas_en_prod = set(
        df_maq_cat[df_maq_cat["ESTADO"].isin(ESTADOS_ACTIVOS)]["ID_MAQUINA"].dropna()
    ) if not df_maq_cat.empty else set()
    maquinas_reportaron = set(df_rep["ID_MAQUINA"].dropna()) if not df_rep.empty else set()
    cumplimiento_pct = min(round(
        len(maquinas_reportaron & maquinas_en_prod) / len(maquinas_en_prod) * 100
        if maquinas_en_prod else 0, 1), 100.0)

    # Semáforo
    if salud_pct >= 90:
        salud_color  = "#10B981"; salud_bg = "#ECFDF5"; salud_borde = "#6EE7B7"
        salud_emoji  = "🟢";     salud_label = "Operación saludable"
        estado_gral  = "VERDE";  estado_color = "#065F46"
        estado_fondo = "#ECFDF5"; estado_borde_css = "#6EE7B7"
        banner_critico = False
    elif salud_pct >= 70:
        salud_color  = "#F59E0B"; salud_bg = "#FFFBEB"; salud_borde = "#FCD34D"
        salud_emoji  = "🟡";     salud_label = "Atención requerida"
        estado_gral  = "AMARILLO"; estado_color = "#92400E"
        estado_fondo = "#FFFBEB"; estado_borde_css = "#FCD34D"
        banner_critico = False
    else:
        salud_color  = "#EF4444"; salud_bg = "#FEF2F2"; salud_borde = "#FCA5A5"
        salud_emoji  = "🔴";     salud_label = "Estado crítico"
        estado_gral  = "ROJO";   estado_color = "#B91C1C"
        estado_fondo = "#FEF2F2"; estado_borde_css = "#FCA5A5"
        banner_critico = True

    # Alertas por categoría — ordenadas por impacto (peor primero)
    sin_rep_df = pd.DataFrame()
    if not alertas.empty and "tipo_alerta" in alertas.columns:
        sin_rep_df = alertas[alertas["tipo_alerta"] == "SIN_REPORTE"].copy()
        if "horas_sin_reporte" in sin_rep_df.columns:
            sin_rep_df = sin_rep_df.sort_values(
                "horas_sin_reporte", ascending=False, na_position="first"
            )

    bajo_rend_df = pd.DataFrame()
    if not maq_df.empty:
        bajo_rend_df = maq_df[
            maq_df["promedio_horas_dia"].notna() & (maq_df["promedio_horas_dia"] < 4)
        ].sort_values("promedio_horas_dia", ascending=True).copy()   # peor rendimiento primero

    comb_inusuales = len(alertas[alertas["tipo_alerta"] == "COMBUSTIBLE_INUSUAL"]) \
                     if not alertas.empty and "tipo_alerta" in alertas.columns else 0

    # Problemas detectados (para mensaje automático)
    problemas = []
    if len(sin_rep_df) > 0:
        problemas.append(f"🔴 {len(sin_rep_df)} máquina(s) sin reporte activo")
    if len(bajo_rend_df) > 0:
        problemas.append(f"🟡 {len(bajo_rend_df)} máquina(s) con bajo rendimiento (<4 hrs/día)")
    if comb_inusuales > 0:
        problemas.append(f"🟡 {comb_inusuales} recarga(s) de combustible fuera de rango")

    # Recomendación automática según estado
    if not problemas:
        recomendacion = "✅ Sin acciones urgentes. Mantener monitoreo regular."
        rec_color = "#065F46"
    elif salud_pct >= 70:
        maq_sin = len(sin_rep_df)
        recomendacion = (
            f"⚡ Contactar a los jefes de obra para verificar el estado de "
            f"{maq_sin} máquina(s) sin reporte. Revisar conectividad de AppSheet."
            if maq_sin > 0 else
            "⚡ Revisar rendimiento de máquinas identificadas y evaluar reasignación de operadores."
        )
        rec_color = "#92400E"
    else:
        recomendacion = (
            f"🚨 ACCIÓN INMEDIATA: {len(sin_rep_df)} máquinas sin reporte. "
            "Activar protocolo de verificación de flota. Contactar supervisores en terreno."
        )
        rec_color = "#B91C1C"

    # Variación respecto al día anterior con flechas
    def delta_con_flecha(col: str) -> dict:
        if act_df.empty or col not in act_df.columns or len(act_df) < 2:
            return {"flecha": "", "texto": "—", "color": "#94A3B8"}
        ordenado = act_df.sort_values("FECHA", ascending=False)
        hoy  = float(ordenado.iloc[0][col])
        ayer = float(ordenado.iloc[1][col])
        dif  = hoy - ayer
        pct  = (dif / ayer * 100) if ayer != 0 else 0
        flecha = "↑" if dif > 0 else ("↓" if dif < 0 else "→")
        color  = "#10B981" if dif > 0 else ("#EF4444" if dif < 0 else "#94A3B8")
        return {
            "flecha": flecha,
            "texto": f"{flecha} {abs(dif):.0f} ({pct:+.1f}%)",
            "color": color,
        }

    delta_maq  = delta_con_flecha("maquinas_activas")
    delta_hrs  = delta_con_flecha("total_horas")

    # ── BANNER CRÍTICO ────────────────────────────────────────────────────────
    if banner_critico:
        st.markdown("""
        <div style="background:#7F1D1D; border-radius:10px; padding:12px 18px;
                    margin-bottom:12px; display:flex; align-items:center; gap:12px">
            <span style="font-size:22px">⛔</span>
            <div>
                <div style="font-size:13px; font-weight:800; color:#FEF2F2; letter-spacing:.04em">
                    RIESGO OPERACIONAL ALTO
                </div>
                <div style="font-size:12px; color:#FECACA; margin-top:2px">
                    La flota está por debajo del umbral mínimo de reporte.
                    Se recomienda revisión inmediata de la operación.
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── MENSAJE DE ESTADO + RECOMENDACIÓN ────────────────────────────────────
    msg_body = "<br>".join(problemas) if problemas \
               else "✅ Sin problemas detectados. Flota operando con normalidad."

    st.markdown(f"""
    <div style="background:{estado_fondo}; border:1.5px solid {estado_borde_css};
                border-left:5px solid {salud_color}; border-radius:10px;
                padding:14px 18px; margin-bottom:14px">
        <div style="display:flex; align-items:center; gap:10px; flex-wrap:wrap;
                    margin-bottom:6px">
            <span style="font-size:13px; font-weight:800; color:{estado_color};
                         letter-spacing:.05em">
                {salud_emoji} ESTADO OPERACIONAL: {estado_gral}
            </span>
            <span style="font-size:11.5px; color:{estado_color}; opacity:.75">
                · {fecha_ini.strftime('%d/%m/%Y')} → {fecha_fin.strftime('%d/%m/%Y')}
            </span>
        </div>
        <div style="font-size:12.5px; color:{estado_color}; opacity:.85;
                    line-height:1.8; margin-bottom:8px">
            {msg_body}
        </div>
        <div style="font-size:12px; color:{rec_color}; font-weight:600;
                    border-top:1px solid {estado_borde_css}; padding-top:8px">
            💡 Recomendación: {recomendacion}
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── HEADER AZUL ──────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#1E40AF 0%,#2563EB 60%,#3B82F6 100%);
                border-radius:14px; padding:22px 28px 18px; margin-bottom:16px;
                position:relative; overflow:hidden">
        <div style="position:absolute;top:-30px;right:-20px;width:180px;height:180px;
                    border-radius:50%;background:rgba(255,255,255,.06)"></div>
        <div style="position:absolute;bottom:-40px;right:140px;width:120px;height:120px;
                    border-radius:50%;background:rgba(255,255,255,.04)"></div>
        <div style="display:flex;align-items:flex-start;justify-content:space-between;
                    flex-wrap:wrap;gap:12px;position:relative;z-index:1">
            <div>
                <div style="font-size:10px;font-weight:700;letter-spacing:.1em;
                            text-transform:uppercase;color:rgba(255,255,255,.55);margin-bottom:4px">
                    HARCHA MAQUINARIA · RESUMEN EJECUTIVO
                </div>
                <div style="font-size:22px;font-weight:800;color:#fff;line-height:1.1;margin-bottom:6px">
                    Vista Ejecutiva de Operaciones
                </div>
                <div style="font-size:12px;color:rgba(255,255,255,.7)">
                    📅 {fecha_ini.strftime('%d %b %Y')} → {fecha_fin.strftime('%d %b %Y')}
                    &nbsp;·&nbsp; {n_dias} días &nbsp;·&nbsp; {len(df_rep):,} reportes
                </div>
            </div>
            <div style="background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.22);
                        border-radius:12px;padding:14px 20px;text-align:center;min-width:148px">
                <div style="font-size:9px;font-weight:700;letter-spacing:.08em;
                            text-transform:uppercase;color:rgba(255,255,255,.5);margin-bottom:3px">
                    SALUD OPERACIONAL
                </div>
                <div style="font-size:32px;font-weight:800;color:#fff;line-height:1">
                    {salud_pct:.0f}%
                </div>
                <div style="font-size:11px;color:rgba(255,255,255,.8);margin-top:3px">
                    {salud_emoji} {salud_label}
                </div>
                <div style="font-size:10px;color:rgba(255,255,255,.5);margin-top:2px">
                    {n_maquinas} de {total_produccion} reportando
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── KPIs (5 + 1 nuevo: cumplimiento) ─────────────────────────────────────
    color_alerta = "kpi-alerta" if n_alertas > 0 else "kpi-verde"
    cumpl_color  = "kpi-verde" if cumplimiento_pct >= 90 \
                   else ("kpi-ambar" if cumplimiento_pct >= 70 else "kpi-alerta")
    kpis = "".join([
        kpi_card("⏱️", "Horas totales",    f"{total_horas:,.0f}",
                 f"prom. {total_horas/n_dias:,.0f} hrs/día" if n_dias else "—"),
        kpi_card("🚜", "Máquinas activas", str(n_maquinas),
                 f"de {total_produccion} en producción"),
        kpi_card("👷", "Operadores",       str(n_operadores),
                 "reportaron actividad", "kpi-verde"),
        kpi_card("⛽", "Litros totales",   f"{total_litros:,.0f}",
                 f"prom. {total_litros/n_dias:,.0f} L/día" if n_dias else "—", "kpi-ambar"),
        kpi_card("📋", "Cumplimiento",     f"{cumplimiento_pct:.0f}%",
                 f"{len(maquinas_reportaron & maquinas_en_prod)} de {len(maquinas_en_prod)} máquinas",
                 cumpl_color),
    ])
    st.markdown(f'<div class="kpi-grid">{kpis}</div>', unsafe_allow_html=True)

    # Deltas con flechas
    cd1, cd2, _, _, _ = st.columns(5)
    if delta_hrs["flecha"]:
        cd1.markdown(
            f"<div style='font-size:12px;color:{delta_hrs['color']};font-weight:600;"
            f"margin-top:-10px;padding-left:4px'>{delta_hrs['texto']} hrs vs ayer</div>",
            unsafe_allow_html=True,
        )
    if delta_maq["flecha"]:
        cd2.markdown(
            f"<div style='font-size:12px;color:{delta_maq['color']};font-weight:600;"
            f"margin-top:-10px;padding-left:4px'>{delta_maq['texto']} máquinas vs ayer</div>",
            unsafe_allow_html=True,
        )

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ── ALERTAS + GRÁFICO ────────────────────────────────────────────────────
    col_izq, col_der = st.columns([2, 3])

    with col_izq:

        # Bloque crítico
        n_crit = len(sin_rep_df)
        st.markdown(f"""
        <div class="seccion" style="border-left:3px solid #EF4444;margin-bottom:12px">
            <div class="seccion-titulo" style="color:#DC2626">
                🚨 Críticas — Sin reporte
                <span class="badge badge-rojo" style="margin-left:6px">{n_crit}</span>
            </div>
        """, unsafe_allow_html=True)

        if sin_rep_df.empty:
            st.markdown("""
            <div style="text-align:center;padding:18px 0">
                <div style="font-size:28px;margin-bottom:5px">✅</div>
                <div style="font-size:12.5px;font-weight:600;color:#065F46">
                    Todas reportando
                </div>
            </div>""", unsafe_allow_html=True)
        else:
            for _, a in sin_rep_df.head(8).iterrows():
                maq  = a.get("MAQUINA", a.get("ID_MAQUINA", "?"))
                tipo = a.get("TIPO_MAQUINA", "")
                hrs  = a.get("horas_sin_reporte", None)
                hrs_txt = f"{hrs:.0f} hs" if pd.notna(hrs) else "sin historial"
                st.markdown(f"""
                <div class="alerta-item alerta-rojo" style="padding:7px 11px">
                    <div class="alerta-dot alerta-dot-r"></div>
                    <div style="flex:1;min-width:0">
                        <div class="alerta-maq" style="overflow:hidden;text-overflow:ellipsis;
                             white-space:nowrap">{str(maq)[:32]}</div>
                        <div class="alerta-desc">{str(tipo)} · {hrs_txt} sin contacto</div>
                    </div>
                </div>""", unsafe_allow_html=True)
            if n_crit > 8:
                st.caption(f"… {n_crit - 8} más en la pestaña Alertas")

        st.markdown("</div>", unsafe_allow_html=True)

        # Bloque atención
        n_br = len(bajo_rend_df)
        st.markdown(f"""
        <div class="seccion" style="border-left:3px solid #F59E0B">
            <div class="seccion-titulo" style="color:#B45309">
                ⚠️ Atención — Bajo rendimiento
                <span class="badge badge-ambar" style="margin-left:6px">{n_br}</span>
            </div>
        """, unsafe_allow_html=True)

        if bajo_rend_df.empty:
            st.markdown("""
            <div style="text-align:center;padding:14px 0">
                <div style="font-size:12.5px;font-weight:600;color:#065F46">
                    ✅ Sin casos
                </div>
            </div>""", unsafe_allow_html=True)
        else:
            for _, row in bajo_rend_df.head(5).iterrows():
                nombre = str(row.get("MAQUINA_TXT", row.get("ID_MAQUINA", "?")))[:30]
                prom   = row["promedio_horas_dia"]
                total_h = row["total_horas"]
                st.markdown(f"""
                <div class="alerta-item alerta-ambar" style="padding:7px 11px">
                    <div class="alerta-dot alerta-dot-a"></div>
                    <div style="flex:1">
                        <div class="alerta-maq">{nombre}</div>
                        <div class="alerta-desc">
                            {prom:.1f} hrs/día · {total_h:.0f} hrs total
                        </div>
                    </div>
                </div>""", unsafe_allow_html=True)
            if n_br > 5:
                st.caption(f"… {n_br - 5} más en Máquinas")

        st.markdown("</div>", unsafe_allow_html=True)

    with col_der:
        st.markdown("""
        <div class="seccion">
            <div class="seccion-titulo">📈 Tendencia operacional</div>
        """, unsafe_allow_html=True)

        metrica_sel = st.radio(
            "",
            ["🚜 Máquinas activas", "⏱ Horas trabajadas", "⛽ Combustible diario"],
            horizontal=True,
            label_visibility="collapsed",
        )

        if not act_df.empty:
            act_plot = act_df.copy()
            act_plot["FECHA"] = pd.to_datetime(act_plot["FECHA"])
            act_plot = act_plot.sort_values("FECHA")

            if metrica_sel == "🚜 Máquinas activas":
                col_g, color_g = "maquinas_activas", "#3B82F6"
                label_g = "Máquinas"
            elif metrica_sel == "⏱ Horas trabajadas":
                col_g, color_g = "total_horas", "#10B981"
                label_g = "Horas"
            else:
                # Combustible diario real desde recargas
                rec = datos.get("recargas", pd.DataFrame())
                if not rec.empty:
                    rec = rec.copy()
                    rec["FECHA_DIA"] = pd.to_datetime(rec["FECHA"]).dt.date
                    comb_dia = (
                        rec.groupby("FECHA_DIA")["LITROS"].sum()
                        .reset_index()
                        .rename(columns={"FECHA_DIA": "FECHA", "LITROS": "litros_dia"})
                    )
                    comb_dia["FECHA"] = pd.to_datetime(comb_dia["FECHA"])
                    act_plot = comb_dia[
                        (comb_dia["FECHA"].dt.date >= fecha_ini) &
                        (comb_dia["FECHA"].dt.date <= fecha_fin)
                    ].sort_values("FECHA")
                    col_g, color_g = "litros_dia", "#F59E0B"
                    label_g = "Litros"
                else:
                    col_g, color_g, label_g = "total_horas", "#10B981", "Horas"

            if col_g in act_plot.columns and not act_plot.empty:
                import plotly.graph_objects as go

                vals = act_plot[col_g]
                avg  = vals.mean()

                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=act_plot["FECHA"],
                    y=vals,
                    name=label_g,
                    marker_color=color_g,
                    marker_line_width=0,
                    opacity=0.85,
                ))
                fig.add_hline(
                    y=avg,
                    line_dash="dot",
                    line_color="#64748B",
                    line_width=1.5,
                    annotation_text=f" Prom: {avg:.1f}",
                    annotation_font_size=11,
                    annotation_font_color="#64748B",
                )
                fig.update_layout(
                    height=250,
                    margin=dict(l=0, r=8, t=8, b=0),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    showlegend=False,
                    xaxis=dict(showgrid=False, tickfont=dict(size=10)),
                    yaxis=dict(gridcolor="#F1F5F9", tickfont=dict(size=10)),
                )
                st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

                cs1, cs2, cs3 = st.columns(3)
                cs1.metric("Mínimo",  f"{vals.min():,.0f}")
                cs2.metric("Promedio",f"{avg:,.1f}")
                cs3.metric("Máximo",  f"{vals.max():,.0f}")

        st.markdown("</div>", unsafe_allow_html=True)

    # ── FILA INFERIOR ─────────────────────────────────────────────────────────
    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    cb1, cb2 = st.columns(2)

    with cb1:
        st.markdown('<div class="seccion"><div class="seccion-titulo">🚜 Top 5 máquinas</div>', unsafe_allow_html=True)
        if not maq_df.empty:
            max_h = maq_df.head(5)["total_horas"].max()
            barras_html = "".join(
                barra(str(row.get("MAQUINA_TXT", row.get("ID_MAQUINA", "?")))[:28],
                      row["total_horas"], max_h, "#3B82F6",
                      f"{row['total_horas']:,.0f} hrs")
                for _, row in maq_df.head(5).iterrows()
            )
            st.markdown(barras_html, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with cb2:
        st.markdown('<div class="seccion"><div class="seccion-titulo">👷 Top 5 operadores</div>', unsafe_allow_html=True)
        if not ops_df.empty:
            max_o    = ops_df.head(5)["total_horas"].max()
            medallas = {1: "🥇", 2: "🥈", 3: "🥉"}
            barras_html = "".join(
                barra(f"{medallas.get(int(row['posicion']),'')} {str(row['USUARIO_TXT'])[:24]}",
                      row["total_horas"], max_o, "#10B981",
                      f"{row['total_horas']:,.0f} hrs")
                for _, row in ops_df.head(5).iterrows()
            )
            st.markdown(barras_html, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)







# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — MÁQUINAS
# ════════════════════════════════════════════════════════════════════════════
#
# ARQUITECTURA:
#   MAQUINAS_BASE (fuente de verdad) → LEFT JOIN ← datos dinámicos del período
#
# REGLA DE ORO: "La lista base manda. Los datos se adaptan a la lista."
#
# Lo que NO cambia nunca (base fija):
#   - MAQUINAS_BASE: 167 máquinas del archivo de mantenciones Harcha
#   - Orden de la tabla (por familia y código)
#   - Qué máquinas se muestran
#
# Lo que cambia con el período seleccionado:
#   - Horas trabajadas
#   - Litros de combustible
#   - L/hr (rendimiento)
#
# Lo que usa historial completo (no depende del período):
#   - Estado (activa/sin actividad: últimos 10 días desde fecha_fin)
#   - Último operador
#   - Ubicación (última obra)
#   - Cambio de obra (🔔)
#   - Días sin reporte
#
# FILTRAR ≠ ELIMINAR:
#   EXT-*, EQUIPO MENOR, etc. se excluyen de ESTA pestaña solamente.
#   El dataset original (datos["maquinas"]) no se modifica.
#
# ════════════════════════════════════════════════════════════════════════════
with tab_maq:

    import re as _re_m

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE A — MAQUINAS_BASE
    # ────────────────────────────────────────────────────────────────────────
    # Fuente: archivo de mantenciones Harcha (hojas c, ca, carg, ex, retro,
    #         bull, tract, motoniv, otro, GE). Descontinuadas excluidas.
    # Case preservado exactamente como aparece en el Excel.
    # Formato de cada tupla:
    #   (codigo_limpio, familia_display, orden_grupo, orden_subgrupo, orden_num)
    # ────────────────────────────────────────────────────────────────────────

    _MAQUINAS_BASE = [
        # ── Camionetas (C-*, MB-*, CUATRI*) ──────────────────────────────
        ("C-02",      "Camioneta",        1,  1,  2),
        ("C-04",      "Camioneta",        1,  1,  4),
        ("C-06",      "Camioneta",        1,  1,  6),
        ("C-09",      "Camioneta",        1,  1,  9),
        ("C-12",      "Camioneta",        1,  1, 12),
        ("C-13",      "Camioneta",        1,  1, 13),
        ("C-14",      "Camioneta",        1,  1, 14),
        ("C-15",      "Camioneta",        1,  1, 15),
        ("C-16",      "Camioneta",        1,  1, 16),
        ("C-17",      "Camioneta",        1,  1, 17),
        ("C-18",      "Camioneta",        1,  1, 18),
        ("C-19",      "Camioneta",        1,  1, 19),
        ("C-20",      "Camioneta",        1,  1, 20),
        ("C-21",      "Camioneta",        1,  1, 21),
        ("C-22",      "Camioneta",        1,  1, 22),
        ("C-23",      "Camioneta",        1,  1, 23),
        ("C-24",      "Camioneta",        1,  1, 24),
        ("C-25",      "Camioneta",        1,  1, 25),
        ("C-26",      "Camioneta",        1,  1, 26),
        ("C-27",      "Camioneta",        1,  1, 27),
        ("C-29",      "Camioneta",        1,  1, 29),
        ("C-30",      "Camioneta",        1,  1, 30),
        ("C-31",      "Camioneta",        1,  1, 31),
        ("C-35",      "Camioneta",        1,  1, 35),
        ("C-36",      "Camioneta",        1,  1, 36),
        ("C-37",      "Camioneta",        1,  1, 37),
        ("C-38",      "Camioneta",        1,  1, 38),
        ("C-39",      "Camioneta",        1,  1, 39),
        ("C-42",      "Camioneta",        1,  1, 42),
        ("C-43",      "Camioneta",        1,  1, 43),
        ("C-45",      "Camioneta",        1,  1, 45),
        ("C-46",      "Camioneta",        1,  1, 46),
        ("C-47",      "Camioneta",        1,  1, 47),
        ("C-48",      "Camioneta",        1,  1, 48),
        ("C-49",      "Camioneta",        1,  1, 49),
        ("C-51",      "Camioneta",        1,  1, 51),
        ("C-52",      "Camioneta",        1,  1, 52),
        ("C-53",      "Camioneta",        1,  1, 53),
        ("C-54",      "Camioneta",        1,  1, 54),
        ("C-55",      "Camioneta",        1,  1, 55),
        ("MB-01",     "Camioneta",        1,  2,  1),
        ("CUATRI 01", "Camioneta",        1,  3,  1),
        ("CUATRI-02", "Camioneta",        1,  3,  2),
        # ── Camiones — por tipo ───────────────────────────────────────────
        ("CA-01",     "Camión Aljibe",   10,  1,  1),
        ("CG-02",     "Camión Ganadero", 10,  2,  2),
        ("CK-03",     "Camión Ganadero", 10,  2,  3),
        ("CM-01",     "Camión Mixer",    10,  3,  1),
        ("CP-02",     "Camión Plano",    10,  4,  2),
        ("CP-03",     "Camión Plano",    10,  4,  3),
        ("CP-04",     "Camión Plano",    10,  4,  4),
        ("CP-05",     "Camión Plano",    10,  4,  5),
        ("CP-06",     "Camión Pluma",    10,  5,  6),
        ("CP-07",     "Camión Pluma",    10,  5,  7),
        ("CP-08",     "Camión Pluma",    10,  5,  8),
        ("CS-01",     "Camión Slurry",   10,  6,  1),
        ("CT-03",     "Camión Tolva",    10,  7,  3),
        ("CT-06",     "Camión Tolva",    10,  7,  6),
        ("CT-07",     "Camión Tolva",    10,  7,  7),
        ("CT-10",     "Camión Tolva",    10,  7, 10),
        ("CT-12",     "Camión Tolva",    10,  7, 12),
        ("CT-13",     "Camión Tolva",    10,  7, 13),
        ("CT-14",     "Camión Tolva",    10,  7, 14),
        ("CT-15",     "Camión Tolva",    10,  7, 15),
        ("CT-16",     "Camión Tolva",    10,  7, 16),
        ("CT-23",     "Camión Tolva",    10,  7, 23),
        ("CT-24",     "Camión Tolva",    10,  7, 24),
        ("CT-25",     "Camión Tolva",    10,  7, 25),
        ("CT-26",     "Camión Tolva",    10,  7, 26),
        ("CT-27",     "Camión Tolva",    10,  7, 27),
        ("CT-28",     "Camión Tolva",    10,  7, 28),
        ("CT-29",     "Camión Tolva",    10,  7, 29),
        ("CT-30",     "Camión Tolva",    10,  7, 30),
        ("CT-31",     "Camión Tolva",    10,  7, 31),
        ("CT-33",     "Camión Tolva",    10,  7, 33),
        ("CT-34",     "Camión Tolva",    10,  7, 34),
        ("CT-35",     "Camión Tolva",    10,  7, 35),
        ("CT-36",     "Camión Tolva",    10,  7, 36),
        ("CT-37",     "Camión Tolva",    10,  7, 37),
        ("CT-38",     "Camión Tolva",    10,  7, 38),
        ("TC-01",     "Tracto Camión",   10,  8,  1),
        ("TC-02",     "Tracto Camión",   10,  8,  2),
        ("TC-03",     "Tracto Camión",   10,  8,  3),
        ("TC-04",     "Tracto Camión",   10,  8,  4),
        ("CI-01",     "Camión Imprimad", 10,  9,  1),
        ("CI-02",     "Camión Imprimad", 10,  9,  2),
        # ── Cargador Frontal / Minicargador ───────────────────────────────
        ("CF-03",     "Cargador Frontal",20,  1,  3),
        ("CF-04",     "Cargador Frontal",20,  1,  4),
        ("CF-05",     "Cargador Frontal",20,  1,  5),
        ("CF-06",     "Cargador Frontal",20,  1,  6),
        ("CF-07",     "Cargador Frontal",20,  1,  7),
        ("CF-08",     "Cargador Frontal",20,  1,  8),
        ("CF-09",     "Cargador Frontal",20,  1,  9),
        ("CF-10",     "Cargador Frontal",20,  1, 10),
        ("MC-01",     "Minicargador",    20,  2,  1),
        ("MC-02",     "Minicargador",    20,  2,  2),
        ("MC-03",     "Minicargador",    20,  2,  3),
        ("MC-04",     "Minicargador",    20,  2,  4),
        # ── Excavadoras ───────────────────────────────────────────────────
        ("EX-03",     "Excavadora",      22,  1,  3),
        ("EX-06",     "Excavadora",      22,  1,  6),
        ("EX-07",     "Excavadora",      22,  1,  7),
        ("EX-08",     "Excavadora",      22,  1,  8),
        ("EX-09",     "Excavadora",      22,  1,  9),
        ("EX-10",     "Excavadora",      22,  1, 10),
        ("EX-11",     "Excavadora",      22,  1, 11),
        ("EX-12",     "Excavadora",      22,  1, 12),
        ("EX-13",     "Excavadora",      22,  1, 13),
        ("EX-14",     "Excavadora",      22,  1, 14),
        ("EX-15",     "Excavadora",      22,  1, 15),
        ("EX-16",     "Excavadora",      22,  1, 16),
        ("EX-17",     "Excavadora",      22,  1, 17),
        ("EX-18",     "Excavadora",      22,  1, 18),
        ("EX-19",     "Excavadora",      22,  1, 19),
        ("EX-20",     "Excavadora",      22,  1, 20),
        ("EX-21",     "Excavadora",      22,  1, 21),
        ("MX-02",     "Mini Excavadora", 22,  2,  2),
        ("MX-03",     "Mini Excavadora", 22,  2,  3),
        # ── Retroexcavadoras ──────────────────────────────────────────────
        ("RX-02",     "Retroexcavadora", 24,  1,  2),
        ("RX-03",     "Retroexcavadora", 24,  1,  3),
        ("RX-04",     "Retroexcavadora", 24,  1,  4),
        ("RX-05",     "Retroexcavadora", 24,  1,  5),
        ("RX-06",     "Retroexcavadora", 24,  1,  6),
        ("RX-07",     "Retroexcavadora", 24,  1,  7),
        ("RX-08",     "Retroexcavadora", 24,  1,  8),
        ("RXm-09",    "Retroexcavadora", 24,  2,  9),  # case exacto del Excel
        # ── Bulldozer ─────────────────────────────────────────────────────
        ("BD-02",     "Bulldozer",       26,  1,  2),
        ("BD-03",     "Bulldozer",       26,  1,  3),
        # ── Tractor ───────────────────────────────────────────────────────
        ("T-01",      "Tractor",         27,  1,  1),
        ("T-02",      "Tractor",         27,  1,  2),
        ("T-03",      "Tractor",         27,  1,  3),
        ("T-04",      "Tractor",         27,  1,  4),
        ("T-05",      "Tractor",         27,  1,  5),
        ("T-06",      "Tractor",         27,  1,  6),
        # ── Motoniveladora ────────────────────────────────────────────────
        ("MN-01",     "Motoniveladora",  28,  1,  1),
        ("MN-02",     "Motoniveladora",  28,  1,  2),
        ("MN-03",     "Motoniveladora",  28,  1,  3),
        ("MN-04",     "Motoniveladora",  28,  1,  4),
        # ── Grúa Móvil ────────────────────────────────────────────────────
        ("GM-01",     "Grúa Móvil",      30,  1,  1),
        ("GM-02",     "Grúa Móvil",      30,  1,  2),
        ("GM-03",     "Grúa Móvil",      30,  1,  3),
        ("GM-04",     "Grúa Móvil",      30,  1,  4),
        # ── Planta / Otros equipos ────────────────────────────────────────
        ("P-01",      "Planta",          30,  2,  1),
        ("P-02",      "Planta",          30,  2,  2),
        ("P-03",      "Planta",          30,  2,  3),
        ("P-04",      "Planta",          30,  2,  4),
        ("P-220",     "Planta",          30,  2,220),
        ("PH-01",     "Otros",           30,  3,  1),
        ("RC-01",     "Rodillo",         30,  4,  1),
        ("RC-02",     "Rodillo",         30,  4,  2),
        ("RL-01",     "Rodillo Liso",    30,  5,  1),
        ("RN-01",     "Rodillo Neumát.", 30,  6,  1),
        ("GV-01",     "Gravilladora",    30,  7,  1),
        ("TE-01",     "Otros",           30,  8,  1),
        ("VH-01",     "Vibrohincador",   30,  9,  1),
        ("BA-01",     "Barredora",       30, 10,  1),
        ("BA-02",     "Barredora",       30, 10,  2),
        ("SE-01",     "Otros",           30, 11,  1),
        # ── Generadores ───────────────────────────────────────────────────
        ("G-01",      "Generador",       50,  1,  1),
        ("G-02",      "Generador",       50,  1,  2),
        ("G-03",      "Generador",       50,  1,  3),
        ("G-04",      "Generador",       50,  1,  4),
        ("G-05",      "Generador",       50,  1,  5),
        ("G-06",      "Generador",       50,  1,  6),
        ("G-07",      "Generador",       50,  1,  7),
        ("G-08",      "Generador",       50,  1,  8),
        ("G-09",      "Generador",       50,  1,  9),
        ("G-10",      "Generador",       50,  1, 10),
        ("G-11",      "Generador",       50,  1, 11),
    ]

    # Set para lookup O(1)
    _BASE_SET = {m[0] for m in _MAQUINAS_BASE}

    # DataFrame de base fija (siempre 167 filas, orden permanente)
    _df_base = pd.DataFrame(
        _MAQUINAS_BASE,
        columns=["CODIGO_LIMPIO","FAMILIA","_og","_os","_on"]
    ).sort_values(["_og","_os","_on"]).reset_index(drop=True)
    _df_base["#"] = range(1, len(_df_base) + 1)

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE B — limpiar_codigo(texto)
    # ────────────────────────────────────────────────────────────────────────
    # Extrae el código limpio de un CODIGO_MAQUINA sucio.
    # PRESERVA el case exacto (no normaliza).
    # Retorna None si no hay código válido.
    #
    # Casos manejados:
    #   "CT-25 WH9819-7"        → "CT-25"
    #   "RXm-09 JOHN DEERE"     → "RXm-09"   (case preservado)
    #   "[CT-14 RCKP70] - FOTON"→ "CT-14"    (corchetes eliminados)
    #   "CUATRI 01"             → "CUATRI 01" (caso especial)
    #   "CUATRI-02"             → "CUATRI-02" (caso especial)
    #   "EXT-03 RETRO..."       → "EXT-03"    (no en base → excluido por filtro)
    #   "EQUIPO MENOR..."       → None         (sin patrón)
    # ────────────────────────────────────────────────────────────────────────

    def limpiar_codigo(texto) -> str | None:
        """
        Extrae el código limpio de un CODIGO_MAQUINA.
        No normaliza el case — respeta el texto tal como viene.
        Solo extrae el patrón LETRAS-NÚMERO del inicio del string.
        """
        if pd.isna(texto) or not str(texto).strip():
            return None

        # Limpiar string: eliminar corchetes iniciales y espacios
        s = _re_m.sub(r"^\[+", "", str(texto).strip()).strip()

        # Caso especial: CUATRI (tiene espacio en lugar de guión: "CUATRI 01")
        if s.upper().startswith("CUATRI"):
            m = _re_m.match(r"(CUATRI[\s\-]\d+)", s, _re_m.IGNORECASE)
            if m:
                # Devolver tal como viene en la base (buscar en BASE_SET)
                candidato = m.group(1)
                # Intentar match directo o normalizado
                for base_cod in _BASE_SET:
                    if base_cod.upper() == candidato.upper():
                        return base_cod  # devuelve el código con case de la base
                return candidato  # si no matchea, devolver como vino
            return None

        # Patrón general: LETRAS (1-4) + "-" + NÚMERO (1-3 dígitos)
        # Preserva el case exacto del input
        m = _re_m.match(r"([A-Za-z]{1,4})-(\d{1,3})", s)
        if m:
            return f"{m.group(1)}-{m.group(2)}"

        return None

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE C — MAPEO: sistema → base
    # ────────────────────────────────────────────────────────────────────────
    # Aplica limpiar_codigo() al dataset original SIN modificarlo.
    # Construye el mapeo CODIGO_LIMPIO → ID_MAQUINA para el LEFT JOIN.
    # ────────────────────────────────────────────────────────────────────────

    # Dataset original intacto — solo leemos, no modificamos
    _df_sistema = datos["maquinas"].copy()

    # Crear columna temporal de código limpio (solo para mapeo)
    _df_sistema["_CODIGO_LIMPIO"] = _df_sistema["CODIGO_MAQUINA"].apply(limpiar_codigo)

    # Filtrar SOLO las que coinciden con la base (no eliminamos del original)
    _df_mapeada = _df_sistema[
        _df_sistema["_CODIGO_LIMPIO"].isin(_BASE_SET)
    ][["ID_MAQUINA","_CODIGO_LIMPIO","EQUIPO_FAMILIA"]].copy()

    # Deduplicar por código limpio (guardar el primero)
    _df_mapeada = _df_mapeada.drop_duplicates(subset=["_CODIGO_LIMPIO"])

    # Mapa rápido: CODIGO_LIMPIO → ID_MAQUINA
    _mapa_id = _df_mapeada.set_index("_CODIGO_LIMPIO")["ID_MAQUINA"].to_dict()

    # Añadir ID_MAQUINA a la base fija (LEFT JOIN: base manda)
    _df_base["ID_MAQUINA"] = _df_base["CODIGO_LIMPIO"].map(_mapa_id)

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE D — DATOS HISTÓRICOS (no dependen del período)
    # ────────────────────────────────────────────────────────────────────────

    _df_rep_hist = datos["reportes"].copy()
    _df_rec_hist = datos.get("recargas", pd.DataFrame()).copy()
    _df_rep_prd  = df_rep.copy()   # ya filtrado por el selector de fechas

    # D1. Máquinas activas en los últimos 10 días (desde fecha_fin)
    _corte_10d = pd.Timestamp(fecha_fin) - pd.Timedelta(days=10)
    _maq_10d   = set()
    if not _df_rep_hist.empty:
        _maq_10d |= set(
            _df_rep_hist[_df_rep_hist["FECHAHORA_INICIO"] >= _corte_10d]
            ["ID_MAQUINA"].dropna()
        )
    if not _df_rec_hist.empty:
        _f_rec = pd.to_datetime(_df_rec_hist["FECHA"], errors="coerce")
        _maq_10d |= set(
            _df_rec_hist[_f_rec >= _corte_10d]["ID_MAQUINA"].dropna()
        )

    # D2. Último operador, fecha y obra por máquina
    _ult: dict = {}
    if not _df_rep_hist.empty:
        _ult = (
            _df_rep_hist.sort_values("FECHAHORA_INICIO")
            .groupby("ID_MAQUINA")
            .agg(
                _op = ("USUARIO_TXT",     "last"),
                _dt = ("FECHAHORA_INICIO","max"),
                _ob = ("OBRA_TXT",        "last"),
            )
            .to_dict("index")
        )

    # D3. Historial de obras para alarma de cambio de obra
    _hist_rep_ob: dict = {}
    if not _df_rep_hist.empty:
        _hist_rep_ob = (
            _df_rep_hist.sort_values("FECHAHORA_INICIO")
            .groupby("ID_MAQUINA")["OBRA_TXT"]
            .apply(lambda s: s.dropna().tolist())
            .to_dict()
        )

    _hist_rec_ob: dict = {}
    if not _df_rec_hist.empty:
        _rec_s = _df_rec_hist.copy()
        _rec_s["_f"] = pd.to_datetime(_rec_s["FECHA"], errors="coerce")
        _hist_rec_ob = (
            _rec_s.sort_values("_f")
            .groupby("ID_MAQUINA")["OBRA_ID"]
            .apply(lambda s: s.dropna().tolist())
            .to_dict()
        )

    # D3b. Historial COMBINADO (reportes + recargas) con timestamp y fuente.
    # ─────────────────────────────────────────────────────────────────────
    # NORMALIZACIÓN DE OBRAS (crítico para eliminar falsos positivos):
    #
    # El campo OBRA_ID de recargas contiene dos tipos de valores:
    #   Tipo A: ID real corto (ej: "11375235", "db040f27") → mapear a nombre
    #   Tipo B: texto directo (ej: "DV04 Choshuenco")       → usar tal cual
    #
    # Ambos se normalizan a upper().strip() + colapso de espacios dobles,
    # de modo que "DV04 Choshuenco" == "DV04 CHOSHUENCO" en la comparación.
    # ─────────────────────────────────────────────────────────────────────

    import re as _re_obra

    # Cargar mapa ID_OBRA → OBRA desde la hoja OBRAS del Excel
    try:
        _df_obras_raw = pd.read_excel(
            st.session_state.archivo_bytes
            if isinstance(st.session_state.get("archivo_bytes"), (bytes, bytearray))
            else open(
                str(__import__("pathlib").Path(__file__).parent
                    / "data" / "raw" / st.session_state.get("archivo_nombre","")),
                "rb"
            ).read()
            if False else io.BytesIO(st.session_state["archivo_bytes"]),
            sheet_name="OBRAS"
        )
        _df_obras_raw = _df_obras_raw.dropna(subset=["ID_OBRA","OBRA"])
    except Exception:
        _df_obras_raw = pd.DataFrame(columns=["ID_OBRA","OBRA"])

    def _es_id_valido(x) -> bool:
        """True si x parece un ID técnico: sin espacios, ≤10 chars, alfanumérico."""
        if pd.isna(x): return False
        s = str(x).strip()
        return (len(s) <= 10
                and " " not in s
                and bool(_re_obra.match(r"^[a-zA-Z0-9]+$", s)))

    # Mapa ID → nombre para IDs válidos solamente
    _mapa_obras = {
        str(r["ID_OBRA"]).strip(): str(r["OBRA"]).strip()
        for _, r in _df_obras_raw.iterrows()
        if _es_id_valido(r["ID_OBRA"]) and pd.notna(r["OBRA"])
    }

    def _norm_obra(texto) -> str:
        """Normaliza nombre de obra: UPPER, strip, colapso de espacios."""
        if pd.isna(texto) or not str(texto).strip():
            return ""
        return _re_obra.sub(r"\s+", " ", str(texto).strip().upper())

    def _obra_de_recarga(obra_id) -> str:
        """
        Convierte OBRA_ID de recarga a nombre normalizado.
        Si es ID válido y está en mapa → usa el nombre real.
        Si es texto directo → normaliza directamente.
        """
        if pd.isna(obra_id): return ""
        s = str(obra_id).strip()
        if _es_id_valido(s) and s in _mapa_obras:
            return _norm_obra(_mapa_obras[s])
        return _norm_obra(s)

    # Construir hist_comb con nombres normalizados
    _hist_comb: dict = {}

    if not _df_rep_hist.empty:
        for _mid, _grp in _df_rep_hist.dropna(
            subset=["OBRA_TXT"]
        ).groupby("ID_MAQUINA"):
            _hist_comb[_mid] = [
                (row["FECHAHORA_INICIO"],
                 _norm_obra(row["OBRA_TXT"]),   # ← normalizado
                 "REP")
                for _, row in _grp[["FECHAHORA_INICIO","OBRA_TXT"]].iterrows()
                if _norm_obra(row["OBRA_TXT"])
            ]

    if not _df_rec_hist.empty:
        _rtmp = _df_rec_hist.copy()
        _rtmp["_f"] = pd.to_datetime(_rtmp["FECHA"], errors="coerce")
        for _mid, _grp in _rtmp.dropna(
            subset=["OBRA_ID","_f"]
        ).groupby("ID_MAQUINA"):
            _evs = _hist_comb.get(_mid, [])
            for _, _row in _grp[["_f","OBRA_ID"]].iterrows():
                _ob = _obra_de_recarga(_row["OBRA_ID"])  # ← normalizado + mapeado
                if _ob:
                    _evs.append((_row["_f"], _ob, "REC"))
            _hist_comb[_mid] = sorted(_evs, key=lambda x: x[0])

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE E — DATOS DINÁMICOS (dependen del período seleccionado)
    # ────────────────────────────────────────────────────────────────────────

    # E1. Horas trabajadas en el período
    _agg_h = pd.DataFrame(columns=["ID_MAQUINA","horas","prom_dia"])
    if not _df_rep_prd.empty:
        _tmp = _df_rep_prd.copy()
        _tmp["_d"] = _tmp["FECHAHORA_INICIO"].dt.date
        _agg_h = (
            _tmp.groupby("ID_MAQUINA")
            .agg(horas=("HORAS_TRABAJADAS","sum"), _dias=("_d","nunique"))
            .reset_index()
        )
        _agg_h["horas"]  = pd.to_numeric(_agg_h["horas"],  errors="coerce").fillna(0)
        _agg_h["_dias"]  = pd.to_numeric(_agg_h["_dias"],  errors="coerce").fillna(0)
        _agg_h["prom_dia"] = (
            _agg_h["horas"] / _agg_h["_dias"].where(_agg_h["_dias"] > 0)
        ).round(1).fillna(0)
        _agg_h = _agg_h[["ID_MAQUINA","horas","prom_dia"]]

    # E2. Litros consumidos en el período
    _agg_l = pd.DataFrame(columns=["ID_MAQUINA","litros"])
    if not _df_rec_hist.empty:
        _rec_p = _df_rec_hist.copy()
        _rec_p["_d"] = pd.to_datetime(_rec_p["FECHA"], errors="coerce").dt.date
        _mask_p = (_rec_p["_d"] >= fecha_ini) & (_rec_p["_d"] <= fecha_fin)
        _agg_l = (
            _rec_p[_mask_p].groupby("ID_MAQUINA")["LITROS"]
            .sum().reset_index().rename(columns={"LITROS":"litros"})
        )
        _agg_l["litros"] = pd.to_numeric(_agg_l["litros"], errors="coerce").fillna(0)

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE F — LEFT JOIN: base fija ← datos dinámicos
    # ────────────────────────────────────────────────────────────────────────
    # La base SIEMPRE tiene 167 filas. El join NUNCA reduce ese número.
    # Máquinas sin datos en el período muestran 0 o "—".

    _tabla = _df_base.copy()
    _tabla = _tabla.merge(_agg_h, on="ID_MAQUINA", how="left")
    _tabla = _tabla.merge(_agg_l, on="ID_MAQUINA", how="left")

    # Forzar dtype numérico (left join puede dejar object)
    for _c in ["horas","prom_dia","litros"]:
        _tabla[_c] = pd.to_numeric(_tabla[_c], errors="coerce").fillna(0)

    # L/hr — rendimiento (0 si no hay datos)
    _tabla["l_hr"] = (
        _tabla["litros"] / _tabla["horas"].where(_tabla["horas"] > 0)
    ).round(2).fillna(0)

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE G — COLUMNAS DE HISTORIAL
    # ────────────────────────────────────────────────────────────────────────

    _ref_ts = pd.Timestamp(fecha_fin)

    # G1. Estado operacional (últimos 10 días)
    def _estado(row) -> str:
        mid = row["ID_MAQUINA"]
        if pd.isna(mid):
            return "Sin datos"       # máquina no encontrada en sistema
        if mid not in _maq_10d:
            return "Sin actividad"
        if row["prom_dia"] > 0 and row["prom_dia"] < 4:
            return "Bajo rend."
        return "Activa"

    _tabla["estado_op"] = _tabla.apply(_estado, axis=1)

    # G2. Último operador
    _tabla["ult_op"] = _tabla["ID_MAQUINA"].apply(
        lambda m: str(_ult.get(m,{}).get("_op","—"))[:32]
                  if pd.notna(m) and _ult.get(m,{}).get("_op") else "—"
    )

    # G3. Ubicación (última obra)
    _tabla["ubicacion"] = _tabla["ID_MAQUINA"].apply(
        lambda m: str(_ult.get(m,{}).get("_ob","Sin datos"))[:35]
                  if pd.notna(m) and _ult.get(m,{}).get("_ob") else "Sin datos"
    )

    # G4. Cambio de obra — lógica mejorada con historial combinado
    # ────────────────────────────────────────────────────────────
    # Aplica SOLO a maquinaria operacional pesada.
    # NO aplica a camionetas, camiones de flete ni generadores.

    _FAMILIAS_SIN_ALARMA = {
        # Camionetas (no hacen reportes operacionales)
        "Camioneta",
        # Camiones de flete / transporte
        "Camión Aljibe", "Camión Ganadero", "Camión Mixer",
        "Camión Plano", "Camión Pluma", "Camión Plano/Pluma",
        "Camión Slurry", "Tracto Camión", "Camión Imprimad",
        # Generadores (equipos fijos)
        "Generador",
    }

    def _cambio_obra_v2(mid, familia) -> tuple:
        """
        Detecta cambio de obra usando historial COMBINADO (reportes + combustible).

        Aplica a: Camión Tolva, Excavadora, Mini Excavadora, Retroexcavadora,
                  Motoniveladora, Bulldozer, Tractor, Cargador Frontal,
                  Minicargador, Grúa Móvil, Rodillo, Gravilladora,
                  Barredora, Planta y Otros.

        Lógica sobre los últimos 5 eventos cronológicos combinados:
          CASO 1: más de 1 obra distinta → ALERTA
          CASO 2: obra actual ≠ obra más frecuente histórica → ALERTA
          CASO 3: última obra de reporte ≠ última obra de recarga → ALERTA

        Retorna: (flag bool, detalle str)
        """
        if pd.isna(mid) or str(familia) in _FAMILIAS_SIN_ALARMA:
            return (False, "")

        eventos = _hist_comb.get(mid, [])
        if not eventos:
            return (False, "")

        # Últimos 5 eventos del historial combinado (ya está ordenado por fecha asc)
        ultimos    = eventos[-5:]
        obras_ult  = [e[1] for e in ultimos]

        # CASO 1: más de 1 obra distinta en los últimos 5 registros
        obras_unicas = list(dict.fromkeys(obras_ult))   # dedup, preserva orden
        if len(obras_unicas) > 1:
            secuencia = " → ".join(o[:20] for o in obras_unicas[:5])
            return (True, f"Cambio detectado: {secuencia}")

        # CASO 2: obra actual ≠ obra más frecuente del historial completo
        from collections import Counter as _Counter
        todas_obras   = [e[1] for e in eventos]
        obra_actual   = obras_ult[-1] if obras_ult else None
        obra_habitual = _Counter(todas_obras).most_common(1)[0][0] \
                        if todas_obras else None
        if obra_actual and obra_habitual and obra_actual != obra_habitual:
            return (True,
                    f"Actual ('{obra_actual[:22]}') ≠ "
                    f"habitual ('{obra_habitual[:22]}')")

        # CASO 3: última obra de reporte ≠ última obra de recarga
        _obras_rep = [e[1] for e in eventos if e[2] == "REP"]
        _obras_rec = [e[1] for e in eventos if e[2] == "REC"]
        if _obras_rep and _obras_rec:
            _ur = _obras_rep[-1].lower().strip()
            _uc = _obras_rec[-1].lower().strip()
            if _ur and _uc and _ur not in _uc and _uc not in _ur:
                return (True,
                        f"Rep. ('{_obras_rep[-1][:22]}') ≠ "
                        f"rec. ('{_obras_rec[-1][:22]}')")

        return (False, "")

    _cambio_res = _tabla.apply(
        lambda r: _cambio_obra_v2(r["ID_MAQUINA"], r["FAMILIA"]), axis=1
    )
    # Convertir flag bool → icono para columna visible, guardar detalle completo
    _tabla["alarma"] = _cambio_res.apply(lambda x: "🔔" if x[0] else "")
    _tabla["motivo"] = _cambio_res.apply(lambda x: x[1])

    # G5. Días sin reporte
    _tabla["dias_sr"] = _tabla["ID_MAQUINA"].apply(
        lambda m: int((_ref_ts - _ult[m]["_dt"]).days)
                  if pd.notna(m) and m in _ult and pd.notna(_ult[m]["_dt"])
                  else None
    )

    # G6. generar_indicadores(maquina) — sistema combinado de iconos
    # ────────────────────────────────────────────────────────────────
    # Retorna:
    #   iconos_str   : string con iconos concatenados, ej: "🔔 💤 📊"
    #   detalle_dict : dict con detalle por indicador
    #
    # Tipos:
    #   🔔 Cambio de obra   — ya calculado en _cambio_res
    #   💤 Inactividad      — sin actividad en > N días (configurable)
    #   📊 Análisis de uso  — solo informativo, NO es alarma
    # ────────────────────────────────────────────────────────────────

    # Umbral configurable de inactividad (días)
    _UMBRAL_INACT = 7

    # Pre-calcular eficiencia por máquina con historial completo de reportes
    # (últimos 30 días desde fecha_fin para análisis de uso)
    _corte_30d = _ref_ts - pd.Timedelta(days=30)
    _efic_map: dict = {}

    if not _df_rep_hist.empty:
        _rep_30d = _df_rep_hist[
            _df_rep_hist["FECHAHORA_INICIO"] >= _corte_30d
        ].copy()
        if not _rep_30d.empty:
            _rep_30d["_dia"] = _rep_30d["FECHAHORA_INICIO"].dt.date
            for _mid_e, _grp_e in _rep_30d.groupby("ID_MAQUINA"):
                _hrs_por_dia = _grp_e.groupby("_dia")["HORAS_TRABAJADAS"].sum()
                _hrs_list    = [h for h in _hrs_por_dia.values if h > 0]
                if not _hrs_list:
                    continue
                import statistics as _stat
                _prom   = round(sum(_hrs_list) / len(_hrs_list), 1)
                _desv   = round(_stat.stdev([float(h) for h in _hrs_list]), 1) if len(_hrs_list) > 1 else 0
                _cv     = (_desv / _prom) if _prom > 0 else 0   # coef. variación

                # Clasificación
                if _prom >= 8:
                    _clf = "Óptimo"
                elif _prom >= 5:
                    _clf = "Normal"
                elif _prom >= 2:
                    _clf = "Bajo"
                else:
                    _clf = "Irregular"

                if _cv > 0.5:
                    _var = "Alta"
                elif _cv > 0.25:
                    _var = "Media"
                else:
                    _var = "Baja"

                # Conclusión automática
                if _clf == "Óptimo" and _var in ("Baja","Media"):
                    _concl = ("✔", "Bien aprovechada")
                elif _clf in ("Bajo","Irregular") or _cv > 0.5:
                    _concl = ("⚠", "Revisar uso")
                else:
                    _concl = ("✔", "Uso normal")

                _efic_map[_mid_e] = {
                    "prom":       _prom,
                    "desv":       _desv,
                    "variacion":  _var,
                    "clasif":     _clf,
                    "dias_datos": len(_hrs_list),
                    "concl":      _concl,
                }

    def _generar_indicadores(row) -> tuple:
        """
        Combina todos los indicadores de una máquina.
        Retorna (iconos_str, detalle_dict).
        iconos_str: "🔔 💤 📊" o subconjunto, o "" si no hay nada.
        """
        mid      = row["ID_MAQUINA"]
        familia  = row["FAMILIA"]
        iconos   = []
        detalle  = {}

        # ── 🔔 Cambio de obra (ya calculado) ─────────────────────────────
        if row["alarma"] == "🔔":
            iconos.append("🔔")
            detalle["cambio_obra"] = row["motivo"]

        # ── 💤 Inactividad ────────────────────────────────────────────────
        # Una máquina activa en el catálogo que lleva > _UMBRAL_INACT días
        # sin ninguna actividad (reporte ni recarga)
        if pd.notna(mid):
            _dias_inact = row["dias_sr"]
            _en_prod    = str(row.get("ESTADO","")).strip() in ESTADOS_ACTIVOS
            if _en_prod and pd.notna(_dias_inact) and int(_dias_inact) > _UMBRAL_INACT:
                iconos.append("💤")
                detalle["inactividad"] = (
                    f"Sin actividad en los últimos {int(_dias_inact)} días"
                )

        # ── 📊 Análisis de uso ────────────────────────────────────────────
        # Solo para maquinaria que tiene datos de historial suficientes
        if pd.notna(mid) and mid in _efic_map:
            _e = _efic_map[mid]
            # Mostrar 📊 solo si hay al menos 5 días de datos
            if _e["dias_datos"] >= 5:
                iconos.append("📊")
                detalle["uso"] = _e

        return (" ".join(iconos), detalle)

    _ind_res = _tabla.apply(_generar_indicadores, axis=1)
    _tabla["indicadores"]      = _ind_res.apply(lambda x: x[0])
    _tabla["indicadores_det"]  = _ind_res.apply(lambda x: x[1])

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE H — KPIs
    # ────────────────────────────────────────────────────────────────────────

    _n_tot   = len(_tabla)
    _n_act   = (_tabla["estado_op"] == "Activa").sum()
    _n_si    = (_tabla["estado_op"] == "Sin actividad").sum()
    _n_br    = (_tabla["estado_op"] == "Bajo rend.").sum()
    _n_sd    = (_tabla["estado_op"] == "Sin datos").sum()
    _n_alarm = (_tabla["alarma"]    == "🔔").sum()

    st.markdown(f"""
    <div style="display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap">
        <div style="background:#065F46;color:#fff;border-radius:8px;
                    padding:10px 20px;text-align:center;min-width:86px">
            <div style="font-size:22px;font-weight:800">{_n_act}</div>
            <div style="font-size:10px;opacity:.85;letter-spacing:.05em">ACTIVAS</div>
        </div>
        <div style="background:#B91C1C;color:#fff;border-radius:8px;
                    padding:10px 20px;text-align:center;min-width:86px">
            <div style="font-size:22px;font-weight:800">{_n_si}</div>
            <div style="font-size:10px;opacity:.85;letter-spacing:.05em">SIN ACTIVIDAD</div>
        </div>
        <div style="background:#92400E;color:#fff;border-radius:8px;
                    padding:10px 20px;text-align:center;min-width:86px">
            <div style="font-size:22px;font-weight:800">{_n_br}</div>
            <div style="font-size:10px;opacity:.85;letter-spacing:.05em">BAJO REND.</div>
        </div>
        <div style="background:#1E40AF;color:#fff;border-radius:8px;
                    padding:10px 20px;text-align:center;min-width:86px">
            <div style="font-size:22px;font-weight:800">{_n_tot}</div>
            <div style="font-size:10px;opacity:.85;letter-spacing:.05em">TOTAL BASE</div>
        </div>
        <div style="background:#6D28D9;color:#fff;border-radius:8px;
                    padding:10px 20px;text-align:center;min-width:86px">
            <div style="font-size:22px;font-weight:800">{_n_alarm}</div>
            <div style="font-size:10px;opacity:.85;letter-spacing:.05em">🔔 CAMBIO OBRA</div>
        </div>
        {'<div style="background:#374151;color:#fff;border-radius:8px;padding:10px 20px;text-align:center;min-width:86px"><div style="font-size:22px;font-weight:800">' + str(_n_sd) + '</div><div style="font-size:10px;opacity:.85;letter-spacing:.05em">SIN DATOS</div></div>' if _n_sd > 0 else ''}
    </div>
    """, unsafe_allow_html=True)

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE I — FILTROS
    # Reducen la vista, NUNCA cambian el orden ni eliminan de la base.
    # ────────────────────────────────────────────────────────────────────────

    st.markdown('<div class="seccion">', unsafe_allow_html=True)
    _fc1, _fc2, _fc3, _fc4 = st.columns([2, 2, 1, 3])

    with _fc1:
        # Familias en orden de aparición (no alfabético)
        _fams_ord = list(dict.fromkeys(_tabla["FAMILIA"].tolist()))
        _fam_sel  = st.selectbox("Familia", ["Todas"] + _fams_ord, key="mf_fam")

    with _fc2:
        _est_ops = ["Todos"] + sorted(_tabla["estado_op"].unique().tolist())
        _est_sel = st.selectbox("Estado", _est_ops, key="mf_est")

    with _fc3:
        _solo_alarm = st.checkbox("🔔 Alarmas", key="mf_alarm")

    with _fc4:
        _q = st.text_input(
            "🔎 Buscar código, familia…",
            placeholder="CT-14, Excavadora, RXm…",
            key="mf_q"
        )

    st.markdown("</div>", unsafe_allow_html=True)

    # Aplicar filtros
    _df_f = _tabla.copy()
    if _fam_sel   != "Todas": _df_f = _df_f[_df_f["FAMILIA"]    == _fam_sel]
    if _est_sel   != "Todos": _df_f = _df_f[_df_f["estado_op"] == _est_sel]
    if _solo_alarm:           _df_f = _df_f[_df_f["indicadores"].str.contains("🔔", na=False)]
    if _q.strip():
        _qn = _q.strip().lower()
        _mask = (
            _df_f["CODIGO_LIMPIO"].str.lower().str.contains(_qn, na=False) |
            _df_f["FAMILIA"].str.lower().str.contains(_qn, na=False)
        )
        _df_f = _df_f[_mask]

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE J — PREPARAR VISTA FINAL
    # ────────────────────────────────────────────────────────────────────────

    # Estado con emoji de color
    _EST_EMOJI = {
        "Activa":       "🟢 Activa",
        "Sin actividad":"🔴 Sin actividad",
        "Bajo rend.":   "🟡 Bajo rend.",
        "Sin datos":    "⚫ Sin datos",
    }

    def _fmt(x, dec=1):
        """Formatea número: 0 o NaN → '—', resto con separador de miles."""
        if pd.isna(x) or x == 0:
            return "—"
        return f"{x:,.{dec}f}"

    _df_v = _df_f[[
        "#","CODIGO_LIMPIO","FAMILIA","estado_op","ult_op",
        "horas","litros","l_hr","ubicacion","indicadores","motivo",
        "indicadores_det","dias_sr"
    ]].copy()

    _df_v["estado_op"]   = _df_v["estado_op"].map(_EST_EMOJI).fillna(_df_v["estado_op"])
    _df_v["horas"]       = _df_v["horas"].apply(lambda x: _fmt(x, 1))
    _df_v["litros"]      = _df_v["litros"].apply(lambda x: _fmt(x, 0))
    _df_v["l_hr"]        = _df_v["l_hr"].apply(lambda x: _fmt(x, 2))
    _df_v["dias_sr"]     = _df_v["dias_sr"].apply(
        lambda x: f"{int(x)} d" if pd.notna(x) else "—"
    )

    _df_v = _df_v.rename(columns={
        "#":               "#",
        "CODIGO_LIMPIO":   "Código",
        "FAMILIA":         "Familia",
        "estado_op":       "Estado",
        "ult_op":          "Último operador",
        "horas":           "Horas",
        "litros":          "Litros",
        "l_hr":            "L/hr",
        "ubicacion":       "Ubicación",
        "indicadores":     "🔔",          # misma columna, ahora con iconos combinados
        "motivo":          "_motivo",      # oculto — usado en modal
        "indicadores_det": "_ind_det",     # oculto — usado en modal
        "dias_sr":         "Días s/rep.",
    })

    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE K — TABLA CON SELECCIÓN DE FILA
    # ────────────────────────────────────────────────────────────────────────
    # Usa on_select="rerun" para detectar la fila seleccionada sin checkboxes.
    # El detalle de la alarma se muestra debajo (NO en popup).
    # La columna "Motivo alarma" se oculta de la vista pero se conserva en _df_v.
    # ────────────────────────────────────────────────────────────────────────

    st.markdown(
        f'<div class="seccion">'
        f'<div class="seccion-titulo">'
        f'Mostrando <b>{len(_df_v)}</b> de <b>{_n_tot}</b> máquinas'
        f'<span style="font-size:11px;color:var(--texto-3);font-weight:400;margin-left:10px">'
        f'Horas/Litros: {fecha_ini.strftime("%d/%m/%Y")} → {fecha_fin.strftime("%d/%m/%Y")}'
        f'  ·  Estado/Alarmas: historial completo'
        f'  ·  <i>Haz clic en una fila para ver detalle</i>'
        f'</span></div>',
        unsafe_allow_html=True,
    )

    # Columnas visibles en la tabla (Motivo alarma se oculta — se muestra en panel)
    _cols_tabla = ["#","Código","Familia","Estado","Último operador",
                   "Horas","Litros","L/hr","Ubicación","🔔","Días s/rep."]

    _evento_sel = st.dataframe(
        _df_v[_cols_tabla],
        use_container_width=True,
        hide_index=True,
        height=min(40 * len(_df_v) + 42, 640),
        on_select="rerun",
        selection_mode="single-row",
        key="maq_sel_tbl",
        column_config={
            "#": st.column_config.NumberColumn("#", width="small", format="%d"),
            "Código": st.column_config.TextColumn(
                "Código", width="small",
                help="Código limpio del catálogo MAQUINAS_BASE."
            ),
            "Familia": st.column_config.TextColumn("Familia"),
            "Estado": st.column_config.TextColumn(
                "Estado",
                help=(
                    "🟢 Activa        — reporte o recarga en últimos 10 días\n"
                    "🔴 Sin actividad — sin movimiento en 10 días\n"
                    "🟡 Bajo rend.    — < 4 hrs/día promedio en el período\n"
                    "⚫ Sin datos     — máquina en base pero no en sistema"
                ),
            ),
            "Último operador": st.column_config.TextColumn("Último operador"),
            "Horas":   st.column_config.TextColumn("Horas",   width="small",
                help="Suma de horas trabajadas en el período seleccionado."),
            "Litros":  st.column_config.TextColumn("Litros",  width="small",
                help="Suma de litros de combustible en el período."),
            "L/hr":    st.column_config.TextColumn("L/hr",    width="small",
                help="Rendimiento: litros / hora trabajada en el período."),
            "Ubicación": st.column_config.TextColumn("Ubicación",
                help="Última obra reportada (historial completo)."),
            "🔔": st.column_config.TextColumn(
                "🔔", width="small",
                help=(
                    "🔔 Cambio de obra detectado\n"
                    "💤 Sin actividad reciente (> 7 días)\n"
                    "📊 Análisis de uso disponible\n"
                    "──\n"
                    "Haz clic en la fila para ver el detalle completo."
                ),
            ),
            "Días s/rep.": st.column_config.TextColumn(
                "Días s/rep.", width="small",
                help="Días desde el último reporte (historial completo)."
            ),
        },
    )

    # ── Modal / tarjeta profesional de detalle ────────────────────────────
    # CRÍTICO: Todo el HTML de la tarjeta se construye en Python como un
    # único string y se pasa a UN SOLO st.markdown(). Si se usan múltiples
    # st.markdown() anidados, Streamlit cierra cada div automáticamente y
    # el HTML aparece como texto plano (el bug visible en la imagen).
    # ─────────────────────────────────────────────────────────────────────
    _filas_sel  = _evento_sel.selection.rows if _evento_sel else []
    _codigo_sel = None

    if _filas_sel:
        _idx_sel    = _filas_sel[0]
        _fila_sel   = _df_v.iloc[_idx_sel]
        _codigo_sel = _fila_sel["Código"]
        _fam_m      = _fila_sel["Familia"]
        _est_m      = _fila_sel["Estado"]
        _op_m       = _fila_sel["Último operador"]
        _hrs_m      = _fila_sel["Horas"]
        _lit_m      = _fila_sel["Litros"]
        _lhr_m      = _fila_sel["L/hr"]
        _ub_m       = _fila_sel["Ubicación"]
        _dia_m      = _fila_sel["Días s/rep."]
        _ind_m      = str(_fila_sel["🔔"])
        _det_m      = _fila_sel.get("_ind_det", {})

        # ── Colores según estado ───────────────────────────────────────────
        if "Activa" in str(_est_m):
            _hdr_bg="#065F46"; _badge_bg="#D1FAE5"; _badge_cl="#065F46"; _badge_txt="ACTIVA"
        elif "Bajo" in str(_est_m):
            _hdr_bg="#92400E"; _badge_bg="#FEF3C7"; _badge_cl="#92400E"; _badge_txt="BAJO REND."
        elif "Sin actividad" in str(_est_m):
            _hdr_bg="#7F1D1D"; _badge_bg="#FEE2E2"; _badge_cl="#B91C1C"; _badge_txt="SIN ACTIVIDAD"
        else:
            _hdr_bg="#374151"; _badge_bg="#F3F4F6"; _badge_cl="#374151"; _badge_txt="SIN DATOS"

        # ── Fechas de último reporte y recarga ────────────────────────────
        _row_data    = _tabla[_tabla["CODIGO_LIMPIO"] == _codigo_sel]
        _id_m        = _row_data["ID_MAQUINA"].iloc[0] if not _row_data.empty else None
        _ult_rep_fmt = "—"
        _ult_rec_fmt = "—"
        if pd.notna(_id_m) and _id_m in _ult:
            _dt = _ult[_id_m].get("_dt")
            if pd.notna(_dt):
                _ult_rep_fmt = pd.Timestamp(_dt).strftime("%d-%m-%Y")
        if pd.notna(_id_m) and not _df_rec_hist.empty:
            _recs_m = _df_rec_hist[_df_rec_hist["ID_MAQUINA"] == _id_m]
            if not _recs_m.empty:
                _ult_rc = pd.to_datetime(_recs_m["FECHA"], errors="coerce").max()
                if pd.notna(_ult_rc):
                    _ult_rec_fmt = pd.Timestamp(_ult_rc).strftime("%d-%m-%Y")

        # ── Construir HTML de indicadores (en Python, no con st.markdown) ─
        _ind_html = ""
        if isinstance(_det_m, dict):

            if "cambio_obra" in _det_m:
                _ind_html += f"""
                <div style="background:#FEF2F2;border-left:3px solid #EF4444;
                            border-radius:0 8px 8px 0;padding:10px 14px;margin-bottom:8px">
                  <div style="font-size:12px;font-weight:700;color:#B91C1C;margin-bottom:3px">
                    🔔 Cambio de obra
                  </div>
                  <div style="font-size:12px;color:#7F1D1D;line-height:1.5">
                    {_det_m['cambio_obra']}
                  </div>
                </div>"""

            if "inactividad" in _det_m:
                _ind_html += f"""
                <div style="background:#F1F5F9;border-left:3px solid #94A3B8;
                            border-radius:0 8px 8px 0;padding:10px 14px;margin-bottom:8px">
                  <div style="font-size:12px;font-weight:700;color:#475569;margin-bottom:3px">
                    💤 Inactividad
                  </div>
                  <div style="font-size:12px;color:#64748B">
                    {_det_m['inactividad']}
                  </div>
                </div>"""

            if "uso" in _det_m:
                _u   = _det_m["uso"]
                _ci  = _u["concl"][0]
                _ct  = _u["concl"][1]
                _cc  = {"✔":"#065F46","⚠":"#92400E","❌":"#B91C1C"}.get(_ci,"#374151")
                _cb  = {"✔":"#D1FAE5","⚠":"#FEF3C7","❌":"#FEE2E2"}.get(_ci,"#F3F4F6")
                _cfc = {"Óptimo":"#065F46","Normal":"#1E40AF",
                        "Bajo":"#92400E","Irregular":"#B91C1C"}.get(_u["clasif"],"#374151")
                _ind_html += f"""
                <div style="background:#F0F9FF;border-left:3px solid #38BDF8;
                            border-radius:0 8px 8px 0;padding:10px 14px;margin-bottom:8px">
                  <div style="font-size:12px;font-weight:700;color:#0369A1;margin-bottom:6px">
                    📊 Uso / Eficiencia — últimos 30 días
                  </div>
                  <div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:6px">
                    <span style="background:#E0F2FE;color:#0369A1;padding:2px 9px;
                                 border-radius:12px;font-size:11px;font-weight:600">
                      {_u['prom']} h/día
                    </span>
                    <span style="background:#E0F2FE;color:#0369A1;padding:2px 9px;
                                 border-radius:12px;font-size:11px;font-weight:600">
                      Var. {_u['variacion']}
                    </span>
                    <span style="background:#E0F2FE;color:{_cfc};padding:2px 9px;
                                 border-radius:12px;font-size:11px;font-weight:600">
                      {_u['clasif']}
                    </span>
                  </div>
                  <div style="background:{_cb};color:{_cc};
                              padding:5px 10px;border-radius:6px;
                              font-size:12px;font-weight:600">
                    {_ci} {_ct}
                  </div>
                </div>"""

        if not _ind_html:
            _ind_html = """
            <div style="text-align:center;padding:28px 0;color:#94A3B8;font-size:13px">
              ✓ Sin indicadores activos
            </div>"""

        # ── Renderizar TODA la tarjeta en un solo st.markdown() ───────────
        st.markdown(f"""
        <div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:14px;
                    overflow:hidden;margin:12px 0 6px;
                    box-shadow:0 4px 20px rgba(0,0,0,.08)">

          <!-- HEADER -->
          <div style="background:{_hdr_bg};padding:16px 20px 14px;
                      display:flex;justify-content:space-between;align-items:flex-start">
            <div>
              <div style="font-size:17px;font-weight:800;color:#fff;
                          letter-spacing:.01em;line-height:1.2;margin-bottom:7px">
                {_codigo_sel} — {_fam_m}
              </div>
              <span style="display:inline-block;padding:3px 11px;
                           background:{_badge_bg};color:{_badge_cl};
                           border-radius:20px;font-size:10.5px;font-weight:700;
                           letter-spacing:.06em">
                {_badge_txt}
              </span>
            </div>
            <div style="color:rgba(255,255,255,.55);font-size:11px;text-align:right;
                        margin-top:2px;line-height:1.6">
              Período analizado<br>
              <span style="color:rgba(255,255,255,.8);font-weight:600">
                {fecha_ini.strftime('%d/%m/%Y')} → {fecha_fin.strftime('%d/%m/%Y')}
              </span>
            </div>
          </div>

          <!-- BODY: 2 columnas -->
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:0">

            <!-- COLUMNA IZQUIERDA: actividad + rendimiento -->
            <div style="padding:18px 20px;border-right:1px solid #E2E8F0">

              <!-- Actividad -->
              <div style="font-size:10.5px;font-weight:700;letter-spacing:.08em;
                          text-transform:uppercase;color:#94A3B8;margin-bottom:10px">
                📡 Actividad
              </div>
              <table style="width:100%;font-size:12.5px;border-collapse:collapse;
                            margin-bottom:16px">
                <tr>
                  <td style="color:#64748B;padding:4px 0;width:50%">Último reporte</td>
                  <td style="font-weight:600;color:#0F172A;text-align:right">
                    {_ult_rep_fmt}
                  </td>
                </tr>
                <tr>
                  <td style="color:#64748B;padding:4px 0">Última recarga</td>
                  <td style="font-weight:600;color:#0F172A;text-align:right">
                    {_ult_rec_fmt}
                  </td>
                </tr>
                <tr>
                  <td style="color:#64748B;padding:4px 0">Días sin reporte</td>
                  <td style="font-weight:600;color:#0F172A;text-align:right">
                    {_dia_m}
                  </td>
                </tr>
                <tr>
                  <td style="color:#64748B;padding:4px 0">Operador</td>
                  <td style="font-weight:600;color:#0F172A;text-align:right;
                              max-width:150px;word-break:break-word">
                    {_op_m}
                  </td>
                </tr>
                <tr>
                  <td style="color:#64748B;padding:4px 0">Ubicación</td>
                  <td style="font-weight:600;color:#0F172A;text-align:right;
                              max-width:150px;word-break:break-word;font-size:11.5px">
                    {_ub_m}
                  </td>
                </tr>
              </table>

              <!-- Rendimiento período -->
              <div style="font-size:10.5px;font-weight:700;letter-spacing:.08em;
                          text-transform:uppercase;color:#94A3B8;margin-bottom:10px">
                ⏱ Rendimiento período
              </div>
              <div style="display:flex;gap:8px">
                <div style="flex:1;background:#EFF6FF;border-radius:10px;
                            padding:11px 8px;text-align:center">
                  <div style="font-size:17px;font-weight:800;color:#1E40AF;
                              line-height:1">{_hrs_m}</div>
                  <div style="font-size:9.5px;color:#64748B;margin-top:4px;
                              letter-spacing:.04em">HORAS</div>
                </div>
                <div style="flex:1;background:#FFFBEB;border-radius:10px;
                            padding:11px 8px;text-align:center">
                  <div style="font-size:17px;font-weight:800;color:#92400E;
                              line-height:1">{_lit_m}</div>
                  <div style="font-size:9.5px;color:#64748B;margin-top:4px;
                              letter-spacing:.04em">LITROS</div>
                </div>
                <div style="flex:1;background:#F0FDF4;border-radius:10px;
                            padding:11px 8px;text-align:center">
                  <div style="font-size:17px;font-weight:800;color:#065F46;
                              line-height:1">{_lhr_m}</div>
                  <div style="font-size:9.5px;color:#64748B;margin-top:4px;
                              letter-spacing:.04em">L/HR</div>
                </div>
              </div>

            </div>

            <!-- COLUMNA DERECHA: indicadores -->
            <div style="padding:18px 20px">
              <div style="font-size:10.5px;font-weight:700;letter-spacing:.08em;
                          text-transform:uppercase;color:#94A3B8;margin-bottom:10px">
                🚨 Indicadores
              </div>
              {_ind_html}
            </div>

          </div>
        </div>
        """, unsafe_allow_html=True)


    # ────────────────────────────────────────────────────────────────────────
    # BLOQUE L — EXPORTACIONES
    # ────────────────────────────────────────────────────────────────────────

    import io as _io_exp
    import openpyxl as _opxl
    from openpyxl.styles import (Font as _Font, PatternFill as _PFill,
                                  Alignment as _Align, Border as _Border,
                                  Side as _Side)

    # ── Helper: escribir DataFrame en hoja con estilos profesionales ────────
    def _df_a_hoja(ws, df, subtitulo=None):
        """DataFrame → hoja openpyxl con cabecera azul, filas alternas y bordes."""
        # Subtítulo opcional
        if subtitulo:
            ws.append([subtitulo])
            ws.cell(ws.max_row, 1).font = _Font(bold=True, size=11, color="1E40AF")
            ws.append([])

        # Cabecera
        ws.append(list(df.columns))
        _hr = ws.max_row
        for _ci, _ in enumerate(df.columns, 1):
            _c = ws.cell(_hr, _ci)
            _c.font      = _Font(bold=True, color="FFFFFF", size=10)
            _c.fill      = _PFill("solid", fgColor="1E40AF")
            _c.alignment = _Align(horizontal="center", vertical="center", wrap_text=True)
            _c.border    = _Border(
                bottom=_Side(style="thin", color="FFFFFF"),
                right=_Side(style="thin", color="FFFFFF"),
            )
        ws.row_dimensions[_hr].height = 22

        # Datos con filas alternas
        _col_alt = "EFF6FF"   # azul muy claro
        for _ri, _row in enumerate(df.itertuples(index=False), 1):
            ws.append(list(_row))
            _dr = ws.max_row
            if _ri % 2 == 0:
                for _ci2 in range(1, len(df.columns) + 1):
                    ws.cell(_dr, _ci2).fill = _PFill("solid", fgColor=_col_alt)
            ws.row_dimensions[_dr].height = 16

        # Anchos automáticos
        for _col in ws.columns:
            _max_l = max(
                (len(str(_c.value or "")) for _c in _col if _c.row > (_hr - 1)),
                default=8
            )
            ws.column_dimensions[_col[0].column_letter].width = min(_max_l + 2, 50)

    # ── A. Exportación general ─────────────────────────────────────────────
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    _col_exp1, _col_exp2 = st.columns([1, 1])

    # DataFrame de export (toma de _df_f — tabla filtrada, datos crudos)
    _df_exp_base = _df_f[[
        "#","CODIGO_LIMPIO","FAMILIA","estado_op","ult_op",
        "horas","litros","l_hr","ubicacion","indicadores","motivo","dias_sr"
    ]].rename(columns={
        "CODIGO_LIMPIO":  "Código",
        "FAMILIA":        "Familia",
        "estado_op":      "Estado",
        "ult_op":         "Último operador",
        "horas":          "Horas período",
        "litros":         "Litros período",
        "l_hr":           "L/hr",
        "ubicacion":      "Ubicación",
        "indicadores":    "Indicadores",
        "motivo":         "Detalle indicadores",
        "dias_sr":        "Días sin reporte",
    })

    with _col_exp1:
        st.download_button(
            "📥 Exportar tabla (CSV)",
            _df_exp_base.to_csv(index=False).encode("utf-8-sig"),
            "maquinas_tabla.csv",
            "text/csv",
            use_container_width=True,
        )

    with _col_exp2:
        # Excel con portada + datos
        _wb_gen  = _opxl.Workbook()
        _ws_port = _wb_gen.active
        _ws_port.title = "Portada"

        # Portada
        _ws_port.append(["HARCHA MAQUINARIA"])
        _ws_port.cell(1, 1).font = _Font(bold=True, size=16, color="1E40AF")
        _ws_port.append(["Reporte de Flota"])
        _ws_port.cell(2, 1).font = _Font(size=13, color="475569")
        _ws_port.append([f"Período: {fecha_ini} → {fecha_fin}"])
        _ws_port.cell(3, 1).font = _Font(size=11)
        _ws_port.append([f"Generado: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"])
        _ws_port.append([])
        _ws_port.append([f"Máquinas mostradas: {len(_df_exp_base)}"])
        _ws_port.append([f"Filtro familia: {_fam_sel}"])
        _ws_port.append([f"Filtro estado: {_est_sel}"])
        _ws_port.column_dimensions["A"].width = 35

        # Hoja datos
        _ws_dat = _wb_gen.create_sheet("Máquinas")
        _df_a_hoja(_ws_dat, _df_exp_base)

        _buf_gen = _io_exp.BytesIO()
        _wb_gen.save(_buf_gen)
        _buf_gen.seek(0)

        st.download_button(
            "📊 Exportar tabla (Excel)",
            _buf_gen.getvalue(),
            "maquinas_tabla.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # ── B. Exportación historial individual ────────────────────────────────
    if _codigo_sel:
        _id_sel = None
        _rows_sel = _df_f[_df_f["CODIGO_LIMPIO"] == _codigo_sel]
        if not _rows_sel.empty:
            _id_sel = _rows_sel["ID_MAQUINA"].iloc[0]

        if _id_sel and pd.notna(_id_sel):
            _row_sel  = _tabla[_tabla["ID_MAQUINA"] == _id_sel].iloc[0]
            _rep_maq  = _df_rep_hist[_df_rep_hist["ID_MAQUINA"] == _id_sel].sort_values(
                "FECHAHORA_INICIO", ascending=False).copy()
            _rec_maq  = _df_rec_hist[_df_rec_hist["ID_MAQUINA"] == _id_sel].copy()
            if not _rec_maq.empty:
                _rec_maq["_f"] = pd.to_datetime(_rec_maq["FECHA"], errors="coerce")
                _rec_maq = _rec_maq.sort_values("_f", ascending=False)

            if not _rep_maq.empty:
                _rep_maq["En período"] = (
                    (_rep_maq["FECHAHORA_INICIO"].dt.date >= fecha_ini) &
                    (_rep_maq["FECHAHORA_INICIO"].dt.date <= fecha_fin)
                )
            if not _rec_maq.empty:
                _rec_maq["En período"] = (
                    (_rec_maq["_f"].dt.date >= fecha_ini) &
                    (_rec_maq["_f"].dt.date <= fecha_fin)
                )

            _eventos_maq = _hist_comb.get(_id_sel, [])
            _df_seq = pd.DataFrame(
                [{"Fecha": e[0], "Obra": e[1], "Fuente": e[2]} for e in _eventos_maq]
            ) if _eventos_maq else pd.DataFrame(columns=["Fecha","Obra","Fuente"])

            # Indicadores de esta máquina
            _ind_row   = _row_sel.get("indicadores", "")
            _ind_det_r = _row_sel.get("indicadores_det", {})
            _det_txt   = []
            if isinstance(_ind_det_r, dict):
                if "cambio_obra" in _ind_det_r:
                    _det_txt.append(f"🔔 {_ind_det_r['cambio_obra']}")
                if "inactividad" in _ind_det_r:
                    _det_txt.append(f"💤 {_ind_det_r['inactividad']}")
                if "uso" in _ind_det_r:
                    _u = _ind_det_r["uso"]
                    _det_txt.append(
                        f"📊 Prom {_u['prom']} h/día · {_u['clasif']} · "
                        f"Variación {_u['variacion']} · {_u['concl'][1]}"
                    )

            # Construir Excel multi-hoja
            _wb_maq = _opxl.Workbook()

            # HOJA 1 — Resumen
            _ws1 = _wb_maq.active
            _ws1.title = "Resumen"
            _ws1.append(["HARCHA MAQUINARIA — FICHA DE MÁQUINA"])
            _ws1.cell(1, 1).font = _Font(bold=True, size=14, color="1E40AF")
            _ws1.append([])

            _seccion_rows = [
                ("IDENTIFICACIÓN",),
                ("Código",     _codigo_sel),
                ("Familia",    _row_sel["FAMILIA"]),
                ("Estado",     _row_sel["estado_op"]),
                ("",),
                ("PERÍODO ANALIZADO",),
                ("Desde",      str(fecha_ini)),
                ("Hasta",      str(fecha_fin)),
                ("",),
                ("RENDIMIENTO PERÍODO",),
                ("Horas",      _row_sel["horas"]),
                ("Litros",     _row_sel["litros"]),
                ("L/hr",       _row_sel["l_hr"]),
                ("",),
                ("HISTORIAL COMPLETO",),
                ("Último operador",  _row_sel["ult_op"]),
                ("Última ubicación", _row_sel["ubicacion"]),
                ("Días sin reporte", _row_sel["dias_sr"] if pd.notna(_row_sel["dias_sr"]) else "Sin historial"),
                ("",),
                ("INDICADORES",),
                ("Iconos activos", _ind_row if _ind_row else "Ninguno"),
            ]
            for _det in _det_txt:
                _seccion_rows.append(("Detalle", _det))

            for _r in _seccion_rows:
                _ws1.append(list(_r))
                _cr = _ws1.max_row
                if len(_r) == 1 and _r[0]:  # títulos de sección
                    _ws1.cell(_cr, 1).font = _Font(bold=True, color="1E40AF", size=10)
                    _ws1.cell(_cr, 1).fill = _PFill("solid", fgColor="EFF6FF")
                elif len(_r) >= 2:
                    _ws1.cell(_cr, 1).font = _Font(bold=True)

            _ws1.column_dimensions["A"].width = 24
            _ws1.column_dimensions["B"].width = 48

            # HOJA 2 — Reportes
            _ws2 = _wb_maq.create_sheet("Reportes")
            if not _rep_maq.empty:
                _cols_r = [c for c in ["FECHAHORA_INICIO","HORAS_TRABAJADAS","USUARIO_TXT",
                                        "OBRA_TXT","DESCRIPCION","En período"] if c in _rep_maq.columns]
                _df_a_hoja(_ws2, _rep_maq[_cols_r].rename(columns={
                    "FECHAHORA_INICIO":"Fecha/Hora","HORAS_TRABAJADAS":"Horas",
                    "USUARIO_TXT":"Operador","OBRA_TXT":"Obra",
                    "DESCRIPCION":"Descripción",
                }))
            else:
                _ws2.append(["Sin reportes en historial"])

            # HOJA 3 — Combustible
            _ws3 = _wb_maq.create_sheet("Combustible")
            if not _rec_maq.empty:
                _cols_c = [c for c in ["_f","LITROS","USUARIO_ID","OBRA_ID",
                                        "ODOMETRO","En período"] if c in _rec_maq.columns]
                _df_a_hoja(_ws3, _rec_maq[_cols_c].rename(columns={
                    "_f":"Fecha","LITROS":"Litros","USUARIO_ID":"Responsable",
                    "OBRA_ID":"Obra","ODOMETRO":"Odómetro",
                }))
            else:
                _ws3.append(["Sin recargas en historial"])

            # HOJA 4 — Análisis de obra
            _ws4 = _wb_maq.create_sheet("Análisis obra")
            if not _df_seq.empty:
                _df_a_hoja(_ws4, _df_seq, f"Secuencia de obras — {_codigo_sel}")
                from collections import Counter as _Cnt
                _conteo = _Cnt(_df_seq["Obra"].tolist()).most_common()
                _ws4.append([])
                _ws4.append(["RESUMEN"])
                _ws4.cell(_ws4.max_row, 1).font = _Font(bold=True, color="1E40AF")
                _ws4.append(["Obra más frecuente:", _conteo[0][0] if _conteo else "—"])
                _ws4.append(["Total eventos:",       len(_df_seq)])
                _ws4.append(["Obras distintas:",     _df_seq["Obra"].nunique()])
            else:
                _ws4.append(["Sin historial de obras"])

            _buf_maq = _io_exp.BytesIO()
            _wb_maq.save(_buf_maq)
            _buf_maq.seek(0)

            st.download_button(
                f"📄 Historial completo — {_codigo_sel}",
                _buf_maq.getvalue(),
                f"historial_{_codigo_sel.replace('-','_').replace(' ','_')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.caption("⚠ Esta máquina no tiene ID vinculado en el sistema.")
    else:
        st.caption("💡 Haz clic en una fila para habilitar la exportación individual.")

# TAB 3 — OPERADORES
# ════════════════════════════════════════════════════════════
with tab_ops:
    if ops_df.empty:
        st.info("Sin datos de operadores para el período seleccionado.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Operadores activos",     n_operadores)
        c2.metric("Horas promedio",          f"{ops_df['total_horas'].mean():,.1f} hrs")
        c3.metric("Operador top",            str(ops_df.iloc[0]["USUARIO_TXT"])[:26])
        c4.metric("Máx horas un operador",   f"{ops_df['total_horas'].max():,.0f} hrs")

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        crank, ctabla = st.columns([2, 3])

        with crank:
            st.markdown('<div class="seccion"><div class="seccion-titulo">🏆 Ranking</div>', unsafe_allow_html=True)
            medallas = {1:"🥇", 2:"🥈", 3:"🥉"}
            rank_html = ""
            for _, row in ops_df.head(15).iterrows():
                pos  = int(row["posicion"])
                med  = medallas.get(pos, f"#{pos}")
                rank_html += f"""
                <div class="rank-row">
                    <span class="rank-pos">{med}</span>
                    <div style="flex:1">
                        <div class="rank-nombre">{str(row['USUARIO_TXT'])[:30]}</div>
                        <div class="rank-sub">{int(row['dias_trabajados'])} días · {int(row['maquinas_distintas'])} máq.</div>
                    </div>
                    <div class="rank-valor">{row['total_horas']:,.0f} hrs</div>
                </div>"""
            st.markdown(rank_html, unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with ctabla:
            st.markdown('<div class="seccion"><div class="seccion-titulo">📊 Horas por operador (top 20)</div>', unsafe_allow_html=True)
            top20_ops = ops_df.head(20).copy()
            top20_ops["Nombre"] = top20_ops["USUARIO_TXT"].str[:26]
            st.bar_chart(top20_ops.set_index("Nombre")["total_horas"],
                         height=360, color="#10B981", horizontal=True, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="seccion"><div class="seccion-titulo">📋 Detalle completo</div>', unsafe_allow_html=True)
        st.dataframe(ops_df.rename(columns={
            "posicion":"#","USUARIO_TXT":"Operador","total_horas":"Horas",
            "dias_trabajados":"Días","total_reportes":"Reportes",
            "maquinas_distintas":"Máquinas","promedio_horas_dia":"Prom hrs/día",
        }), use_container_width=True, hide_index=True, height=280)
        st.download_button("⬇ Exportar CSV",
            ops_df.to_csv(index=False).encode("utf-8-sig"),
            "operadores.csv", "text/csv")
        st.markdown("</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# TAB 4 — COMBUSTIBLE
# ════════════════════════════════════════════════════════════
with tab_comb:
    if comb_df.empty:
        st.info("Sin datos de combustible para el período seleccionado.")
    else:
        ncol = "MAQUINA" if "MAQUINA" in comb_df.columns else "ID_MAQUINA"
        top_c = str(comb_df.iloc[0].get("MAQUINA", comb_df.iloc[0]["ID_MAQUINA"]))

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total litros",           f"{total_litros:,.0f} L")
        c2.metric("Máquinas con recargas",  len(comb_df))
        c3.metric("Promedio por recarga",   f"{comb_df['promedio_litros_recarga'].mean():,.1f} L")
        c4.metric("Mayor consumidora",      top_c[:22] if pd.notna(top_c) else "—")

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        cg, cp = st.columns([3, 2])

        with cg:
            st.markdown('<div class="seccion"><div class="seccion-titulo">⛽ Consumo por máquina (top 20)</div>', unsafe_allow_html=True)
            top20_c = comb_df.head(20).copy()
            top20_c["Nombre"] = top20_c[ncol].fillna(top20_c["ID_MAQUINA"]).str[:28]
            st.bar_chart(top20_c.set_index("Nombre")["total_litros"],
                         height=360, color="#F59E0B", horizontal=True, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with cp:
            st.markdown('<div class="seccion"><div class="seccion-titulo">📊 Participación en consumo</div>', unsafe_allow_html=True)
            max_l = comb_df.head(10)["total_litros"].max()
            barras_html = ""
            for _, row in comb_df.head(10).iterrows():
                nombre = str(row.get("MAQUINA", row["ID_MAQUINA"]))[:22]
                lts    = row["total_litros"]
                pct    = (lts / total_litros * 100) if total_litros > 0 else 0
                barras_html += barra(nombre, lts, max_l, "#F59E0B",
                                     f"{lts:,.0f} L ({pct:.1f}%)")
            st.markdown(barras_html, unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="seccion"><div class="seccion-titulo">📋 Detalle de recargas</div>', unsafe_allow_html=True)
        cols_c = [c for c in [ncol,"TIPO_MAQUINA","total_litros",
                               "num_recargas","promedio_litros_recarga","ultima_recarga"]
                  if c in comb_df.columns]
        st.dataframe(comb_df[cols_c].rename(columns={
            ncol:"Máquina","TIPO_MAQUINA":"Tipo",
            "total_litros":"Litros","num_recargas":"Recargas",
            "promedio_litros_recarga":"Prom. L/recarga","ultima_recarga":"Última recarga",
        }), use_container_width=True, hide_index=True, height=280)
        st.download_button("⬇ Exportar CSV",
            comb_df.to_csv(index=False).encode("utf-8-sig"),
            "combustible.csv", "text/csv")
        st.markdown("</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# TAB 5 — ALERTAS
# ════════════════════════════════════════════════════════════
with tab_alertas:
    if alertas.empty:
        st.markdown("""
        <div style="text-align:center; padding:56px 0">
            <div style="font-size:44px; margin-bottom:10px">✅</div>
            <div style="font-size:20px; font-weight:700; color:#065F46">Sin alertas activas</div>
            <div style="font-size:13px; color:#64748B; margin-top:6px">
                Todas las máquinas están reportando correctamente
            </div>
        </div>""", unsafe_allow_html=True)
    else:
        tipos_al = alertas["tipo_alerta"].unique() if "tipo_alerta" in alertas.columns else []
        cols_ka  = st.columns(len(tipos_al) + 1)
        cols_ka[0].metric("Total alertas", n_alertas)
        for i, tipo in enumerate(tipos_al):
            n = len(alertas[alertas["tipo_alerta"] == tipo])
            cols_ka[i+1].metric(tipo.replace("_"," ").title(), n)

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

        tipo_sel = st.selectbox("Filtrar por tipo",
            ["Todas las alertas"] + list(tipos_al), label_visibility="collapsed")
        df_al = alertas if tipo_sel == "Todas las alertas" \
                else alertas[alertas["tipo_alerta"] == tipo_sel]

        # Cards de máquinas sin reporte
        sin_rep = df_al[df_al.get("tipo_alerta", pd.Series(dtype=str)) == "SIN_REPORTE"] \
                  if "tipo_alerta" in df_al.columns else pd.DataFrame()

        if not sin_rep.empty:
            st.markdown('<div class="seccion"><div class="seccion-titulo" style="color:#DC2626">🚨 Máquinas sin reporte</div>', unsafe_allow_html=True)
            cols_c = st.columns(3)
            for i, (_, a) in enumerate(sin_rep.head(12).iterrows()):
                maq  = a.get("MAQUINA", a.get("ID_MAQUINA", "?"))
                tipo = a.get("TIPO_MAQUINA", "")
                desc = a.get("descripcion", "")
                hrs  = a.get("horas_sin_reporte", None)
                hrs_html = (f'<div class="alerta-desc" style="font-weight:700;color:#DC2626">'
                            f'{hrs:.0f} horas de silencio</div>' if pd.notna(hrs) else "")
                with cols_c[i % 3]:
                    st.markdown(f"""
                    <div class="alerta-item alerta-rojo" style="flex-direction:column; gap:5px">
                        <div style="display:flex; align-items:center; gap:7px">
                            <div class="alerta-dot alerta-dot-r"></div>
                            <span class="badge badge-rojo">SIN REPORTE</span>
                        </div>
                        <div class="alerta-maq">{str(maq)[:30]}</div>
                        <div class="alerta-desc">{str(tipo)} · {desc}</div>
                        {hrs_html}
                    </div>""", unsafe_allow_html=True)
            if len(sin_rep) > 12:
                st.caption(f"Mostrando 12 de {len(sin_rep)} alertas")
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="seccion"><div class="seccion-titulo">📋 Todas las alertas</div>', unsafe_allow_html=True)
        cols_al = [c for c in ["tipo_alerta","ID_MAQUINA","MAQUINA","TIPO_MAQUINA",
                                "descripcion","ultimo_reporte","horas_sin_reporte",
                                "FECHA","LITROS"] if c in df_al.columns]
        st.dataframe(df_al[cols_al].rename(columns={
            "tipo_alerta":"Tipo","ID_MAQUINA":"ID Máquina","MAQUINA":"Máquina",
            "TIPO_MAQUINA":"Tipo equipo","descripcion":"Descripción",
            "ultimo_reporte":"Último reporte","horas_sin_reporte":"Horas sin reporte",
            "FECHA":"Fecha","LITROS":"Litros",
        }), use_container_width=True, hide_index=True, height=340)
        st.download_button("⬇ Exportar alertas CSV",
            df_al.to_csv(index=False).encode("utf-8-sig"),
            "alertas.csv", "text/csv")
        st.markdown("</div>", unsafe_allow_html=True)


# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style="margin-top:28px; padding-top:14px; border-top:1px solid #E2E8F0;
            display:flex; justify-content:space-between; align-items:center">
    <span style="font-size:11.5px; color:#94A3B8">
        🚜 Harcha Maquinaria · Monitor Operacional v1.1
    </span>
    <span style="font-size:11.5px; color:#94A3B8">
        {fecha_ini.strftime('%d/%m/%Y')} → {fecha_fin.strftime('%d/%m/%Y')}
        · {len(df_rep):,} reportes procesados
    </span>
</div>
""", unsafe_allow_html=True)
